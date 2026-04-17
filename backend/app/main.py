"""
Food Delivery Forecast Studio — FastAPI Backend
"""
import json
import os
import sqlite3
import uuid
from datetime import datetime
from typing import List, Optional

from fastapi import FastAPI, HTTPException, Response
from fastapi.middleware.cors import CORSMiddleware

from .models.schemas import (
    ForecastRequest,
    ForecastResult,
    SaveForecastResponse,
    ForecastListItem,
)
from .models.forecast_engine import run_forecast
from .excel.generator import generate_excel

# ── App setup ──────────────────────────────────────────────────────────────────

app = FastAPI(
    title="Food Delivery Forecast Studio",
    version="1.0.0",
    description="Backend API for forecasting food delivery operations.",
)

_ON_VERCEL = bool(os.environ.get("VERCEL"))

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"] if _ON_VERCEL else ["http://localhost:3000", "http://localhost:5173", "http://127.0.0.1:3000", "http://127.0.0.1:5173"],
    allow_credentials=not _ON_VERCEL,  # credentials + wildcard origin is not allowed
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── SQLite setup ───────────────────────────────────────────────────────────────

# On Vercel the filesystem is read-only except /tmp; persistence is ephemeral per-instance.
DB_PATH = "/tmp/forecasts.db" if _ON_VERCEL else os.path.join(os.path.dirname(__file__), "..", "forecasts.db")


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS forecasts (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL,
            config_json TEXT NOT NULL,
            results_json TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


@app.on_event("startup")
async def startup_event():
    init_db()


# ── Routes ─────────────────────────────────────────────────────────────────────

@app.get("/health")
async def health():
    return {"status": "ok", "version": "1.0.0"}


@app.post("/api/forecast/calculate", response_model=ForecastResult)
async def calculate_forecast(request: ForecastRequest):
    """
    Run the combined forecast model.
    Returns immediately without saving — use /save to persist.
    """
    try:
        result = run_forecast(request)
        return result
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/forecast/save", response_model=SaveForecastResponse)
async def save_forecast(request: ForecastRequest):
    """Run forecast and save to SQLite."""
    try:
        result = run_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    forecast_id = str(uuid.uuid4())
    created_at = datetime.utcnow().isoformat() + "Z"

    conn = get_db()
    conn.execute(
        "INSERT INTO forecasts (id, name, created_at, config_json, results_json) VALUES (?, ?, ?, ?, ?)",
        (
            forecast_id,
            request.name,
            created_at,
            request.model_dump_json(),
            result.model_dump_json(),
        ),
    )
    conn.commit()
    conn.close()

    return SaveForecastResponse(id=forecast_id, name=request.name, created_at=created_at)


@app.get("/api/forecast/list", response_model=List[ForecastListItem])
async def list_forecasts():
    """List all saved forecasts (metadata only)."""
    conn = get_db()
    rows = conn.execute(
        "SELECT id, name, created_at, config_json, results_json FROM forecasts ORDER BY created_at DESC"
    ).fetchall()
    conn.close()

    items = []
    for row in rows:
        try:
            config = json.loads(row["config_json"])
            results = json.loads(row["results_json"])
            summary = results.get("summary", {})
            items.append(
                ForecastListItem(
                    id=row["id"],
                    name=row["name"],
                    created_at=row["created_at"],
                    horizon_weeks=config.get("horizon_weeks", 0),
                    country=config.get("country", ""),
                    models_active=config.get("models_active", []),
                    total_orders=summary.get("total_orders", 0),
                )
            )
        except Exception:
            continue

    return items


@app.get("/api/forecast/{forecast_id}", response_model=ForecastResult)
async def get_forecast(forecast_id: str):
    """Retrieve a saved forecast by ID."""
    conn = get_db()
    row = conn.execute(
        "SELECT results_json FROM forecasts WHERE id = ?", (forecast_id,)
    ).fetchone()
    conn.close()

    if not row:
        raise HTTPException(status_code=404, detail="Forecast not found")

    data = json.loads(row["results_json"])
    return ForecastResult(**data)


@app.delete("/api/forecast/{forecast_id}")
async def delete_forecast(forecast_id: str):
    """Delete a saved forecast."""
    conn = get_db()
    result = conn.execute("DELETE FROM forecasts WHERE id = ?", (forecast_id,))
    conn.commit()
    conn.close()
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="Forecast not found")
    return {"deleted": forecast_id}


@app.post("/api/forecast/export-excel")
async def export_excel(request: ForecastRequest):
    """Run forecast and return an Excel file for download."""
    try:
        result = run_forecast(request)
        excel_bytes = generate_excel(result, request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    filename = f"forecast_{request.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/benchmarks")
async def get_benchmarks():
    """Return benchmark data for UI display."""
    benchmarks_path = os.path.join(os.path.dirname(__file__), "data", "benchmarks.json")
    try:
        with open(benchmarks_path, "r") as f:
            return json.load(f)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Benchmarks file not found")


@app.get("/api/holidays/{country}")
async def get_holidays(country: str):
    """Return holiday data for a specific country."""
    holidays_path = os.path.join(os.path.dirname(__file__), "data", "holidays.json")
    try:
        with open(holidays_path, "r") as f:
            all_holidays = json.load(f)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Holidays file not found")

    country_data = all_holidays.get(country.upper())
    if not country_data:
        raise HTTPException(status_code=404, detail=f"No holiday data for country: {country}")

    return country_data


@app.post("/api/forecast/scenarios")
async def run_scenarios(request: ForecastRequest):
    """
    Run B2 scenario analysis: base / upside / downside + tornado sensitivity.
    """
    # Ensure B2 is in models_active
    if "B2" not in request.models_active:
        request.models_active = list(request.models_active) + ["B2"]

    try:
        result = run_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    return {
        "scenarios": result.scenarios.model_dump() if result.scenarios else None,
        "sensitivity_tornado": [t.model_dump() for t in result.sensitivity_tornado] if result.sensitivity_tornado else [],
    }


# ── Markov v3 Routes ───────────────────────────────────────────────────────────

from .models.markov_schemas import MarkovForecastRequest, MarkovForecastResult
from .models.markov_engine import run_markov_forecast


@app.post("/api/markov/calculate", response_model=MarkovForecastResult)
async def markov_calculate(request: MarkovForecastRequest):
    """Run the Markov + Funnel v3 forecast model."""
    try:
        result = run_markov_forecast(request)
        return result
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/markov/save", response_model=SaveForecastResponse)
async def markov_save(request: MarkovForecastRequest):
    """Run Markov forecast and save to SQLite."""
    try:
        result = run_markov_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    forecast_id = str(uuid.uuid4())
    created_at = datetime.utcnow().isoformat() + "Z"
    conn = get_db()
    conn.execute(
        "INSERT INTO forecasts (id, name, created_at, config_json, results_json) VALUES (?, ?, ?, ?, ?)",
        (forecast_id, request.name, created_at, request.model_dump_json(), result.model_dump_json()),
    )
    conn.commit()
    conn.close()
    return SaveForecastResponse(id=forecast_id, name=request.name, created_at=created_at)


@app.get("/api/markov/assumptions-packs")
async def get_assumptions_packs():
    """List available assumptions packs."""
    import glob
    packs_dir = os.path.join(os.path.dirname(__file__), "data", "assumptions_packs")
    packs = []
    for path in glob.glob(os.path.join(packs_dir, "*.json")):
        with open(path) as f:
            p = json.load(f)
            packs.append({
                "id": os.path.basename(path).replace(".json", ""),
                "name": p.get("pack_name", ""),
                "description": p.get("pack_description", ""),
            })
    return packs


@app.get("/api/markov/assumptions-packs/{pack_id}")
async def get_assumptions_pack(pack_id: str):
    """Get a specific assumptions pack."""
    path = os.path.join(os.path.dirname(__file__), "data", "assumptions_packs", f"{pack_id}.json")
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Pack '{pack_id}' not found")
    with open(path) as f:
        return json.load(f)


@app.post("/api/markov/validate-matrix")
async def validate_transition_matrix(data: dict):
    """Validate that each row of a transition matrix sums to 1.0."""
    matrix = data.get("matrix", [])
    results = []
    for i, row in enumerate(matrix):
        s = sum(row)
        results.append({"row": i, "sum": round(s, 4), "valid": abs(s - 1.0) <= 0.001})
    return {"rows": results, "all_valid": all(r["valid"] for r in results)}


from .excel.markov_generator import generate_markov_excel


@app.post("/api/markov/export-excel")
async def markov_export_excel(request: MarkovForecastRequest):
    """Run Markov forecast and return MBB-grade Excel."""
    try:
        result = run_markov_forecast(request)
        excel_bytes = generate_markov_excel(result, request, palette_key=request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    filename = f"markov_{request.name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.post("/api/markov/export-pptx")
async def markov_export_pptx(raw: _FRequest):
    """Run Markov forecast and return a 5-slide dark-theme PPTX."""
    from .pptx.generic_generator import generate_pptx_generic as _gen_pptx

    try:
        body = await raw.json()
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid JSON body")

    insights = body.pop('_insights', [])
    palette  = body.get('palette', 'navy')

    try:
        req_obj = MarkovForecastRequest(**body)
        raw_result = run_markov_forecast(req_obj)
        result = raw_result if isinstance(raw_result, dict) else raw_result.model_dump()
        config = req_obj.model_dump()
        pptx_bytes = _gen_pptx(result, config, 'D1', 'Markov Cohort Engine', 'D', insights, palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    filename = f"d1_markov_{datetime.now().strftime('%Y%m%d')}.pptx"
    return Response(
        content=pptx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Wizard Model Routes (D2-P5) ────────────────────────────────────────────────

from .models.d2_cohort_engine import CohortRequest, run_cohort_forecast
from .models.d3_funnel_engine import FunnelRequest, run_funnel_forecast
from .models.d4_frequency_engine import FrequencyRequest, run_frequency_forecast
from .models.d5_winback_engine import WinbackRequest, run_winback_forecast
from .models.s1_onboarding_engine import OnboardingRequest, run_onboarding_forecast
from .models.s2_portfolio_engine import PortfolioRequest, run_portfolio_forecast
from .models.s3_engagement_engine import EngagementRequest, run_engagement_forecast
from .models.s4_health_engine import S4HealthRequest, run_health_forecast
from .models.p1_network_engine import NetworkRequest, run_network_forecast
from .models.p2_incrementality_engine import IncrementalityRequest, run_incrementality_forecast
from .models.p3_delivery_engine import DeliveryRequest, run_delivery_forecast
from .models.p4_competitive_engine import CompetitiveRequest, run_competitive_forecast
from .models.p5_equilibrium_engine import EquilibriumRequest, run_equilibrium_forecast


@app.post("/api/models/d2/calculate")
async def d2_calculate(request: CohortRequest):
    try:
        return run_cohort_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/d3/calculate")
async def d3_calculate(request: FunnelRequest):
    try:
        return run_funnel_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/d4/calculate")
async def d4_calculate(request: FrequencyRequest):
    try:
        return run_frequency_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/d5/calculate")
async def d5_calculate(request: WinbackRequest):
    try:
        return run_winback_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/s1/calculate")
async def s1_calculate(request: OnboardingRequest):
    try:
        return run_onboarding_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/s2/calculate")
async def s2_calculate(request: PortfolioRequest):
    try:
        return run_portfolio_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/s3/calculate")
async def s3_calculate(request: EngagementRequest):
    try:
        return run_engagement_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/s4/calculate")
async def s4_calculate(request: S4HealthRequest):
    try:
        return run_health_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/p1/calculate")
async def p1_calculate(request: NetworkRequest):
    try:
        return run_network_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/p2/calculate")
async def p2_calculate(request: IncrementalityRequest):
    try:
        return run_incrementality_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/p3/calculate")
async def p3_calculate(request: DeliveryRequest):
    try:
        return run_delivery_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/p4/calculate")
async def p4_calculate(request: CompetitiveRequest):
    try:
        return run_competitive_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


@app.post("/api/models/p5/calculate")
async def p5_calculate(request: EquilibriumRequest):
    try:
        return run_equilibrium_forecast(request)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))


# ── Excel Export Routes (D2–P5) ────────────────────────────────────────────────

from .excel.generic_generator import generate_generic_excel


@app.post("/api/models/d2/export-excel")
async def d2_export_excel(request: CohortRequest):
    try:
        result = run_cohort_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "D2 — Cohort Retention & LTV", "D2", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"d2_cohort_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/d3/export-excel")
async def d3_export_excel(request: FunnelRequest):
    try:
        result = run_funnel_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "D3 — Funnel Conversion", "D3", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"d3_funnel_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/d4/export-excel")
async def d4_export_excel(request: FrequencyRequest):
    try:
        result = run_frequency_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "D4 — Frequency & Wallet Share", "D4", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"d4_frequency_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/d5/export-excel")
async def d5_export_excel(request: WinbackRequest):
    try:
        result = run_winback_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "D5 — Reactivation & Winback", "D5", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"d5_winback_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/s1/export-excel")
async def s1_export_excel(request: OnboardingRequest):
    try:
        result = run_onboarding_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "S1 — Restaurant Onboarding & Maturation", "S1", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"s1_onboarding_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/s2/export-excel")
async def s2_export_excel(request: PortfolioRequest):
    try:
        result = run_portfolio_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "S2 — Portfolio & Selection Effect", "S2", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"s2_portfolio_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/s3/export-excel")
async def s3_export_excel(request: EngagementRequest):
    try:
        result = run_engagement_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "S3 — Restaurant Engagement & Performance", "S3", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"s3_engagement_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/s4/export-excel")
async def s4_export_excel(request: S4HealthRequest):
    try:
        result = run_health_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "S4 — Restaurant Health Score", "S4", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"s4_health_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/p1/export-excel")
async def p1_export_excel(request: NetworkRequest):
    try:
        result = run_network_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "P1 — Network Effects & Liquidity", "P1", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"p1_network_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/p2/export-excel")
async def p2_export_excel(request: IncrementalityRequest):
    try:
        result = run_incrementality_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "P2 — Incrementality & Cannibalization", "P2", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"p2_incrementality_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/p3/export-excel")
async def p3_export_excel(request: DeliveryRequest):
    try:
        result = run_delivery_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "P3 — Delivery Economics & Capacity", "P3", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"p3_delivery_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/p4/export-excel")
async def p4_export_excel(request: CompetitiveRequest):
    try:
        result = run_competitive_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "P4 — Competitive Dynamics", "P4", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"p4_competitive_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@app.post("/api/models/p5/export-excel")
async def p5_export_excel(request: EquilibriumRequest):
    try:
        result = run_equilibrium_forecast(request)
        excel_bytes = generate_generic_excel(result if isinstance(result, dict) else result.model_dump(), "P5 — Marketplace Equilibrium", "P5", request.model_dump(), request.palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))
    filename = f"p5_equilibrium_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(content=excel_bytes, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f'attachment; filename="{filename}"'})


# ── PPTX Export Routes (D2–P5) ────────────────────────────────────────────────

from fastapi import Request as _FRequest
from .pptx.generic_generator import generate_pptx_generic as _gen_pptx

_MODEL_META = {
    'd2': ('D2', 'Cohort Retention & LTV',             'D', CohortRequest,         run_cohort_forecast),
    'd3': ('D3', 'Funnel Conversion',                   'D', FunnelRequest,          run_funnel_forecast),
    'd4': ('D4', 'Frequency & Wallet Share',            'D', FrequencyRequest,       run_frequency_forecast),
    'd5': ('D5', 'Reactivation & Winback',              'D', WinbackRequest,         run_winback_forecast),
    's1': ('S1', 'Restaurant Onboarding & Maturation',  'S', OnboardingRequest,      run_onboarding_forecast),
    's2': ('S2', 'Portfolio & Selection Effect',        'S', PortfolioRequest,       run_portfolio_forecast),
    's3': ('S3', 'Restaurant Engagement',               'S', EngagementRequest,      run_engagement_forecast),
    's4': ('S4', 'Restaurant Health Score',             'S', S4HealthRequest,        run_health_forecast),
    'p1': ('P1', 'Network Effects & Liquidity',         'P', NetworkRequest,         run_network_forecast),
    'p2': ('P2', 'Incrementality & Cannibalization',    'P', IncrementalityRequest,  run_incrementality_forecast),
    'p3': ('P3', 'Delivery Economics & Capacity',       'P', DeliveryRequest,        run_delivery_forecast),
    'p4': ('P4', 'Competitive Dynamics',                'P', CompetitiveRequest,     run_competitive_forecast),
    'p5': ('P5', 'Marketplace Equilibrium',             'P', EquilibriumRequest,     run_equilibrium_forecast),
}


@app.post("/api/models/{model_id}/export-pptx")
async def export_pptx(model_id: str, request: _FRequest):
    mid = model_id.lower()
    if mid not in _MODEL_META:
        raise HTTPException(status_code=404, detail=f"Model '{model_id}' not found")

    model_code, model_name, perspective, RequestClass, engine_fn = _MODEL_META[mid]

    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid JSON body")

    insights = body.pop('_insights', [])
    palette  = body.get('palette', 'navy')

    try:
        req_obj = RequestClass(**body)
        raw = engine_fn(req_obj)
        result = raw if isinstance(raw, dict) else raw.model_dump()
        config = req_obj.model_dump()
        pptx_bytes = _gen_pptx(result, config, model_code, model_name, perspective, insights, palette)
    except Exception as e:
        raise HTTPException(status_code=422, detail=str(e))

    filename = f"{mid}_{model_code.lower()}_{datetime.now().strftime('%Y%m%d')}.pptx"
    return Response(
        content=pptx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Feedback Endpoint ──────────────────────────────────────────────────────────

from pydantic import BaseModel as _BM
from typing import Optional as _Opt

class FeedbackPayload(_BM):
    nps: int
    message: str = ""
    email: _Opt[str] = None
    company: _Opt[str] = None

@app.post("/api/feedback")
async def submit_feedback(payload: FeedbackPayload):
    try:
        import resend
        resend.api_key = os.environ.get("RESEND_API_KEY", "")
        if not resend.api_key:
            raise HTTPException(status_code=503, detail="Servicio de email no configurado")

        nps_label = "Promotor 😊" if payload.nps >= 9 else "Neutro 😐" if payload.nps >= 7 else "Detractor 😞"
        body_lines = [
            f"<h2>Nuevo Feedback — Forecast Studio</h2>",
            f"<p><strong>NPS:</strong> {payload.nps}/10 — {nps_label}</p>",
        ]
        if payload.message:
            body_lines.append(f"<p><strong>Mensaje:</strong><br>{payload.message.replace(chr(10), '<br>')}</p>")
        if payload.email:
            body_lines.append(f"<p><strong>Email:</strong> {payload.email}</p>")
        if payload.company:
            body_lines.append(f"<p><strong>Empresa:</strong> {payload.company}</p>")
        body_lines.append(f"<p style='color:#888;font-size:12px'>Enviado: {datetime.now().strftime('%d/%m/%Y %H:%M')}</p>")

        params = {
            "from": "Forecast Studio <onboarding@resend.dev>",
            "to": ["tusalario.io@gmail.com"],
            "subject": f"[Forecast Studio] NPS {payload.nps}/10 — {'Nuevo feedback'}",
            "html": "\n".join(body_lines),
        }
        resend.Emails.send(params)
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── Sensitivity / Tornado Analysis ────────────────────────────────────────────

_SENSITIVITY_MAP = {
    'd2': (CohortRequest,      run_cohort_forecast,      'total_revenue'),
    'd3': (FunnelRequest,       run_funnel_forecast,      'total_orders'),
    'd4': (FrequencyRequest,    run_frequency_forecast,   'total_orders'),
    'd5': (WinbackRequest,      run_winback_forecast,     'total_orders'),
    's1': (OnboardingRequest,   run_onboarding_forecast,  'total_orders'),
    's2': (PortfolioRequest,    run_portfolio_forecast,   'total_revenue'),
    's3': (EngagementRequest,   run_engagement_forecast,  'total_orders'),
    's4': (S4HealthRequest,     run_health_forecast,      'revenue_at_risk'),
    'p1': (NetworkRequest,      run_network_forecast,     'total_orders'),
    'p2': (IncrementalityRequest, run_incrementality_forecast, 'blended_roi'),
    'p3': (DeliveryRequest,     run_delivery_forecast,    'total_orders'),
    'p4': (CompetitiveRequest,  run_competitive_forecast, 'total_orders'),
    'p5': (EquilibriumRequest,  run_equilibrium_forecast, 'total_revenue'),
}

_SENSITIVITY_LABELS = {
    'horizon_weeks':        'Horizonte (semanas)',
    'aov':                  'AOV (valor por orden)',
    'take_rate':            'Take Rate',
    'weekly_new_users':     'Nuevos usuarios/semana',
    'uplift_observed_pct':  'Uplift observado',
    'organic_baseline':     'Baseline orgánico',
    'promoted_orders_per_week': 'Órdenes promovidas/sem',
    'orders_per_active_per_week': 'Órdenes/usuario activo',
    'churn_rate':           'Tasa de churn',
    'reactivation_rate':    'Tasa de reactivación',
    'dormant_base':         'Base dormida',
    'weekly_new_restaurants': 'Nuevos restaurantes/sem',
    'avg_orders_per_restaurant': 'Órdenes/restaurante',
    'fleet_size':           'Tamaño de flota',
    'orders_per_courier_per_week': 'Órdenes/courier/sem',
    'market_size_weekly_orders': 'Tamaño de mercado',
    'your_share_pct':       'Market share actual',
    'competitor_share_pct': 'Share de competidores',
}

@app.post("/api/models/{model_id}/sensitivity")
async def model_sensitivity(model_id: str, raw: _FRequest):
    mid = model_id.lower()
    if mid not in _SENSITIVITY_MAP:
        raise HTTPException(status_code=404, detail=f"Modelo {model_id} no encontrado")

    RequestClass, engine_fn, kpi_key = _SENSITIVITY_MAP[mid]
    try:
        body = await raw.json()
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid JSON body")

    # Base run
    try:
        base_req = RequestClass(**body)
        base_result = engine_fn(base_req)
        if not isinstance(base_result, dict):
            base_result = base_result.model_dump()
        base_kpi = float(base_result.get('summary', {}).get(kpi_key, 0) or 0)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Error en modelo base: {e}")

    SKIP_KEYS = {'palette', 'currency', 'country'}
    results = []

    for field, value in body.items():
        if field in SKIP_KEYS:
            continue
        if not isinstance(value, (int, float)) or isinstance(value, bool):
            continue
        if value == 0:
            continue

        row = {'param': field, 'label': _SENSITIVITY_LABELS.get(field, field.replace('_', ' ').title()), 'base_value': value}
        for label, delta in [('low', 0.8), ('high', 1.2)]:
            modified = dict(body)
            modified[field] = value * delta
            try:
                mod_req = RequestClass(**modified)
                mod_result = engine_fn(mod_req)
                if not isinstance(mod_result, dict):
                    mod_result = mod_result.model_dump()
                mod_kpi = float(mod_result.get('summary', {}).get(kpi_key, 0) or 0)
                pct = (mod_kpi - base_kpi) / abs(base_kpi) * 100 if base_kpi != 0 else 0
                row[f'kpi_{label}'] = round(mod_kpi, 2)
                row[f'pct_{label}'] = round(pct, 1)
            except Exception:
                row[f'kpi_{label}'] = base_kpi
                row[f'pct_{label}'] = 0.0
        results.append(row)

    # Sort by max absolute impact, take top 5
    results.sort(key=lambda r: max(abs(r.get('pct_low', 0)), abs(r.get('pct_high', 0))), reverse=True)
    results = results[:5]

    return {
        'model_id': mid,
        'kpi_key': kpi_key,
        'base_kpi': round(base_kpi, 2),
        'tornado': results,
    }


# ── Combined Forecast Excel Export ────────────────────────────────────────────

from .excel.generic_generator import generate_excel_generic as _gen_excel


@app.post("/api/combined/export-excel")
async def combined_export_excel(request: _FRequest):
    """
    Generate a multi-tab MBB Excel for a combined forecast run.
    Body: { global_config: {...}, models: [{ model_id, model_name, config, result }] }
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=422, detail="Invalid JSON body")

    global_config = body.get("global_config", {})
    models_data = body.get("models", [])

    if not models_data:
        raise HTTPException(status_code=422, detail="No model results provided")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import io
        from datetime import datetime

        def _fill(hex_color):
            return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

        def _font(color="FFFFFF", size=11, bold=True):
            return Font(name="Calibri", size=size, bold=bold, color=color)

        def _border(style="thin", color="CCCCCC"):
            s = Side(border_style=style, color=color)
            return Border(left=s, right=s, top=s, bottom=s)

        wb = Workbook()

        # ── Summary tab ──
        ws_sum = wb.active
        ws_sum.title = "Resumen Combinado"

        currency = global_config.get("currency", "MXN")
        horizon = global_config.get("horizon_weeks", 12)
        aov = global_config.get("aov", 290)
        take_rate = global_config.get("take_rate", 0.22)

        # Banner
        ws_sum.merge_cells("A1:F1")
        ws_sum["A1"] = "COMBINED FORECAST — RESUMEN EJECUTIVO"
        ws_sum["A1"].font = _font("FFFFFF", 14, True)
        ws_sum["A1"].fill = _fill("1e3a5f")
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 32

        ws_sum.merge_cells("A2:F2")
        ws_sum["A2"] = f"Horizonte: {horizon} semanas | AOV: {currency} {aov} | Take Rate: {take_rate*100:.0f}% | {datetime.utcnow().strftime('%Y-%m-%d')}"
        ws_sum["A2"].font = _font("94a3b8", 10, False)
        ws_sum["A2"].fill = _fill("0f172a")
        ws_sum["A2"].alignment = Alignment(horizontal="center")
        ws_sum.row_dimensions[2].height = 20

        # Beta disclaimer
        ws_sum.merge_cells("A3:F3")
        ws_sum["A3"] = "BETA: Outputs son independientes por modelo — no suman un P&L unificado."
        ws_sum["A3"].font = _font("d97706", 9, False)
        ws_sum["A3"].fill = _fill("1c1008")
        ws_sum["A3"].alignment = Alignment(horizontal="center")

        # Headers
        headers = ["Modelo", "Nombre", "Tipo", f"Revenue ({currency})", "Órdenes", "KPI adicional"]
        PERSPECTIVE_COLORS_XL = {"D": "1e3a8a", "S": "14532d", "P": "3b0764"}
        for ci, h in enumerate(headers, 1):
            c = ws_sum.cell(row=5, column=ci, value=h)
            c.font = _font("FFFFFF", 10, True)
            c.fill = _fill("1e293b")
            c.border = _border()
            c.alignment = Alignment(horizontal="center")
        ws_sum.row_dimensions[5].height = 18

        # Model rows
        REVENUE_KEYS = {
            "D2": "total_revenue", "D3": "total_revenue", "D4": "total_revenue", "D5": "total_revenue",
            "S1": "total_revenue", "S2": "total_revenue", "S3": "incremental_revenue", "S4": "revenue_at_risk",
            "P1": "total_revenue", "P2": "net_contribution", "P3": "total_revenue",
            "P4": "revenue_vs_baseline", "P5": "total_revenue",
        }
        ORDERS_KEYS = {
            "D3": "total_orders", "D4": "total_orders", "D5": "total_orders",
            "S1": "total_orders", "S2": "total_orders", "S3": "incremental_orders_total",
            "P1": "total_orders", "P2": "total_incremental_orders", "P3": "total_orders",
            "P4": None, "P5": "total_orders",
        }
        TYPE_LABELS = {
            "D2": "Baseline", "D3": "Baseline", "D4": "Incremental", "D5": "Incremental",
            "S1": "Baseline", "S2": "Baseline", "S3": "Incremental", "S4": "Riesgo",
            "P1": "Baseline", "P2": "Incremental", "P3": "Baseline", "P4": "Competitivo", "P5": "Equilibrio",
        }
        PERSPECTIVE_MAP = {
            "D2": "D", "D3": "D", "D4": "D", "D5": "D",
            "S1": "S", "S2": "S", "S3": "S", "S4": "S",
            "P1": "P", "P2": "P", "P3": "P", "P4": "P", "P5": "P",
        }

        for ri, md in enumerate(models_data, 6):
            mid = md.get("model_id", "").upper()
            mname = md.get("model_name", mid)
            summary = (md.get("result") or {}).get("summary", {})

            rev_key = REVENUE_KEYS.get(mid, "total_revenue")
            ord_key = ORDERS_KEYS.get(mid)
            rev_val = summary.get(rev_key, 0) or 0
            ord_val = summary.get(ord_key, "") if ord_key else ""
            type_label = TYPE_LABELS.get(mid, "—")
            persp = PERSPECTIVE_MAP.get(mid, "D")
            persp_color = PERSPECTIVE_COLORS_XL.get(persp, "1e293b")

            # Extra KPI: first unused numeric summary key
            extra_kpi = ""
            used = {rev_key, ord_key}
            for k, v in summary.items():
                if k not in used and isinstance(v, (int, float)):
                    extra_kpi = f"{k}: {v:.2f}"
                    break

            row_data = [mid, mname, type_label, round(rev_val, 0), ord_val, extra_kpi]
            for ci, val in enumerate(row_data, 1):
                c = ws_sum.cell(row=ri, column=ci, value=val)
                c.border = _border("thin", "334155")
                c.alignment = Alignment(horizontal="center" if ci != 2 else "left", vertical="center")
                if ci == 1:
                    c.font = _font("e2e8f0", 9, True)
                    c.fill = _fill(persp_color)
                elif ci == 4 and isinstance(val, (int, float)):
                    is_risk = type_label == "Riesgo"
                    c.font = Font(name="Calibri", size=10, bold=True,
                                  color="ef4444" if is_risk else "10b981")
                    c.number_format = f'"{currency}" #,##0'
                else:
                    c.font = Font(name="Calibri", size=9, color="cbd5e1")
                    c.fill = _fill("0f172a" if ri % 2 == 0 else "111827")

        # Col widths
        for col, width in [(1, 8), (2, 30), (3, 14), (4, 18), (5, 14), (6, 30)]:
            ws_sum.column_dimensions[get_column_letter(col)].width = width

        # ── One tab per model ──
        for md in models_data:
            mid = md.get("model_id", "").upper()
            mname = md.get("model_name", mid)
            result = md.get("result") or {}
            summary = result.get("summary", {})
            weekly = result.get("weekly", [])

            tab_title = f"{mid}"[:31]
            ws = wb.create_sheet(title=tab_title)

            # Tab header
            ws.merge_cells("A1:D1")
            ws["A1"] = f"{mid} — {mname}"
            ws["A1"].font = _font("FFFFFF", 12, True)
            ws["A1"].fill = _fill(PERSPECTIVE_COLORS_XL.get(PERSPECTIVE_MAP.get(mid, "D"), "1e293b"))
            ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[1].height = 28

            # Summary section
            ws["A3"] = "RESUMEN"
            ws["A3"].font = _font("94a3b8", 9, True)
            for ri2, (k, v) in enumerate(summary.items(), 4):
                ws.cell(row=ri2, column=1, value=k.replace("_", " ").title()).font = Font(name="Calibri", size=9, color="94a3b8")
                vc = ws.cell(row=ri2, column=2, value=round(v, 4) if isinstance(v, float) else v)
                vc.font = Font(name="Calibri", size=9, bold=True, color="e2e8f0")

            # Weekly data section
            if weekly:
                start_row = 4 + len(summary) + 2
                ws.cell(row=start_row, column=1, value="DATOS SEMANALES").font = _font("94a3b8", 9, True)
                start_row += 1

                if weekly:
                    headers_weekly = list(weekly[0].keys())
                    for ci, h in enumerate(headers_weekly, 1):
                        c = ws.cell(row=start_row, column=ci, value=h)
                        c.font = _font("FFFFFF", 9, True)
                        c.fill = _fill("1e293b")
                        c.alignment = Alignment(horizontal="center")

                    for wi, week_row in enumerate(weekly, start_row + 1):
                        for ci, h in enumerate(headers_weekly, 1):
                            v = week_row.get(h, "")
                            vc = ws.cell(row=wi, column=ci, value=round(v, 4) if isinstance(v, float) else v)
                            vc.font = Font(name="Calibri", size=9, color="cbd5e1")
                            vc.fill = PatternFill(start_color="0f172a" if wi % 2 == 0 else "111827",
                                                  end_color="0f172a" if wi % 2 == 0 else "111827",
                                                  fill_type="solid")

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

        # ── Stream out ──
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=combined_forecast.xlsx"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")
