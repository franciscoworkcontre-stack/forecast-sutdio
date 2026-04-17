"""
Generic MBB-grade Excel generator for all wizard models (D2-P5).
Generates a 3-tab workbook:
  1. Resumen Ejecutivo  — CEO-ready headline dashboard (MBB style)
  2. Escenarios          — Bear / Base / Bull side-by-side comparison
  3. Datos Detallados    — Full weekly/period data table
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

from .theme import get_palette


# ── Style helpers ──────────────────────────────────────────────────────────────

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(color: str = "FFFFFF", size: int = 11, bold: bool = True, italic: bool = False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, italic=italic, color=color)


def _border(style: str = "thin", color: str = "CCCCCC") -> Border:
    s = Side(border_style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _auto_width(ws, max_width: int = 48):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


def _write_header_row(ws, row: int, values: list, palette: dict, height: int = 22):
    primary = palette['primary']
    font_color = palette['font_header']
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill = _fill(primary)
        cell.font = _font(color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border("thin", "888888")
    ws.row_dimensions[row].height = height


def _write_section_title(ws, row: int, title: str, col_span: int, palette: dict):
    secondary = palette['secondary']
    font_color = palette['font_header']
    cell = ws.cell(row=row, column=1, value=title)
    cell.fill = _fill(secondary)
    cell.font = _font(color=font_color, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    ws.row_dimensions[row].height = 18


def _fmt_value(v, pct_hint: bool = False) -> str:
    if isinstance(v, bool):
        return "Sí" if v else "No"
    if isinstance(v, float):
        if pct_hint:
            return f"{v:.1f}%"
        if abs(v) < 0.01:
            return f"{v:.4f}"
        if abs(v) < 10:
            return f"{v:.2f}"
        return f"{v:,.0f}"
    if isinstance(v, int):
        return f"{v:,}"
    return str(v) if v is not None else "—"


_KPI_LABELS = {
    'total_revenue':                  'Revenue Total',
    'total_orders':                   'Órdenes Totales',
    'total_weekly_cost':              'Costo Semanal Total',
    'total_weekly_incremental_revenue': 'Revenue Incremental / Sem',
    'blended_roi':                    'ROI Combinado',
    'blended_cannibalization_rate_pct': 'Tasa Canibalización',
    'avg_ltv_cac':                    'LTV / CAC Promedio',
    'best_channel':                   'Mejor Canal',
    'breakeven_weekly_orders':        'Breakeven (órdenes/sem)',
    'biggest_drop_step':              'Paso con Mayor Caída',
    'overall_conversion_rate':        'Conversión Total Funnel',
    'at_risk_count':                  'Restaurantes en Riesgo',
    'revenue_at_risk':                'Revenue en Riesgo',
    'final_phase':                    'Fase de Liquidez',
    'final_share_pct':                'Market Share Final',
    'total_restaurants_added':        'Restaurantes Activados',
    'bottleneck_week':                'Semana Cuello de Botella',
    'roi':                            'ROI del Programa',
    'avg_health_score':               'Score de Salud Promedio',
    'uplift_pct':                     'Uplift Portfolio',
    'is_sustainable':                 'Sostenibilidad del Negocio',
    'horizon_weeks':                  'Horizonte (semanas)',
    'ltv_cac_ratio':                  'LTV / CAC',
}

_KPI_UNITS = {
    'blended_roi': '%',
    'blended_cannibalization_rate_pct': '%',
    'overall_conversion_rate': '%',
    'final_share_pct': '%',
    'uplift_pct': '%',
    'avg_health_score': '/100',
}


def _kpi_label(k: str) -> str:
    return _KPI_LABELS.get(k, k.replace('_', ' ').title())


def _kpi_unit(k: str) -> str:
    return _KPI_UNITS.get(k, '')


# ── Tab 1: Resumen Ejecutivo (CEO-ready) ───────────────────────────────────────

def _build_executive_tab(wb, result: dict, model_name: str, model_id: str, config: dict, palette: dict):
    ws = wb.create_sheet("Resumen Ejecutivo")
    ws.sheet_view.showGridLines = False

    primary   = palette['primary']
    secondary = palette['secondary']
    light     = palette['light']
    accent    = palette['accent']
    fh        = palette['font_header']

    # ── Banner ──
    row = 1
    ws.merge_cells(f"A{row}:H{row}")
    c = ws.cell(row=row, column=1, value=f"FORECAST STUDIO  |  {model_id} — {model_name.upper()}")
    c.fill = _fill(primary)
    c.font = _font(color=fh, size=14, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 34
    row += 1

    ws.merge_cells(f"A{row}:H{row}")
    sub = ws.cell(row=row, column=1, value=(
        f"Generado: {datetime.now().strftime('%d %b %Y  %H:%M')}  "
        f"| País: {config.get('country','—')}  "
        f"| Horizonte: {config.get('horizon_weeks','—')} semanas  "
        f"| {config.get('currency','MXN')} {config.get('aov','—')} AOV"
    ))
    sub.fill = _fill(secondary)
    sub.font = _font(color=fh, size=9, bold=False)
    sub.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16
    row += 2

    # ── Headline KPIs (metric cards in a grid) ──
    summary = result.get('summary', {})
    numeric_kpis = [(k, v) for k, v in summary.items()
                    if isinstance(v, (int, float)) and not isinstance(v, bool)]
    text_kpis    = [(k, v) for k, v in summary.items()
                    if isinstance(v, str) or isinstance(v, bool)]

    if numeric_kpis:
        ws.merge_cells(f"A{row}:H{row}")
        lbl = ws.cell(row=row, column=1, value="INDICADORES CLAVE")
        lbl.fill = _fill(secondary)
        lbl.font = _font(color=fh, size=9)
        lbl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1

        # Header row: Indicador | Valor | Unidad | Escenario Bear | Escenario Bull
        _write_header_row(ws, row, [
            "Indicador", "Valor Base", "Unidad",
            f"Bear (×0.6)", f"Bull (×1.4)", "Variación Bear↔Bull"
        ], palette, height=24)
        row += 1

        for i, (k, v) in enumerate(numeric_kpis):
            bg = light if i % 2 == 0 else "FFFFFF"
            is_pct = 'pct' in k or 'rate' in k
            unit = _kpi_unit(k)

            bear_val = v * 0.6
            bull_val = v * 1.4
            range_pct = ((bull_val - bear_val) / abs(v)) * 100 if v != 0 else 0

            row_data = [
                _kpi_label(k),
                v,
                unit or ('MXN' if 'revenue' in k or 'cost' in k else ''),
                bear_val,
                bull_val,
                f"{range_pct:+.0f}%",
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _fill(bg)
                if col == 1:
                    cell.font = _font(color="222222", size=10, bold=True)
                elif col == 2:
                    cell.font = _font(color="1a3a5c", size=11, bold=True)
                elif col == 4:
                    cell.font = _font(color="991b1b", size=10, bold=False)  # bear red
                elif col == 5:
                    cell.font = _font(color="065f46", size=10, bold=False)  # bull green
                elif col == 6:
                    cell.font = _font(color="374151", size=10, bold=False, italic=True)
                else:
                    cell.font = _font(color="374151", size=10, bold=False)
                cell.alignment = Alignment(horizontal="center" if col > 1 else "left",
                                           vertical="center", indent=1 if col == 1 else 0)
                if isinstance(val, float) and col in (2, 4, 5):
                    cell.number_format = '#,##0.00' if abs(val) < 10 else '#,##0'
            ws.row_dimensions[row].height = 18
            row += 1
        row += 1

    # ── Text / Categorical KPIs ──
    if text_kpis:
        ws.merge_cells(f"A{row}:H{row}")
        lbl2 = ws.cell(row=row, column=1, value="DIAGNÓSTICO CUALITATIVO")
        lbl2.fill = _fill(secondary)
        lbl2.font = _font(color=fh, size=9)
        lbl2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1
        _write_header_row(ws, row, ["Indicador", "Valor"], palette, height=20)
        row += 1
        for i, (k, v) in enumerate(text_kpis):
            bg = light if i % 2 == 0 else "FFFFFF"
            kc = ws.cell(row=row, column=1, value=_kpi_label(k))
            kc.fill = _fill(bg); kc.font = _font(color="222222", size=10, bold=True)
            kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc = ws.cell(row=row, column=2, value=_fmt_value(v))
            vc.fill = _fill(bg); vc.font = _font(color="1a3a5c", size=10, bold=False)
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1

    # ── Config params ──
    skip_keys = {'palette', 'channels', 'segments', 'campaigns', 'buckets', 'tiers',
                 'events', 'restaurant_types', 'zone_configs', 'signals', 'retention_curve'}
    config_items = [(k, v) for k, v in config.items()
                    if k not in skip_keys and not isinstance(v, (list, dict))]
    if config_items:
        ws.merge_cells(f"A{row}:H{row}")
        lbl3 = ws.cell(row=row, column=1, value="SUPUESTOS Y PARÁMETROS")
        lbl3.fill = _fill(secondary)
        lbl3.font = _font(color=fh, size=9)
        lbl3.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1
        _write_header_row(ws, row, ["Parámetro", "Valor", "Descripción"], palette, height=20)
        row += 1
        param_descs = {
            'horizon_weeks': 'Semanas de proyección',
            'aov': 'Valor promedio por transacción',
            'take_rate': 'Comisión de la plataforma (0–1)',
            'currency': 'Moneda de reporte',
            'country': 'País de operación',
        }
        for i, (k, v) in enumerate(config_items):
            bg = light if i % 2 == 0 else "FFFFFF"
            kc = ws.cell(row=row, column=1, value=k.replace('_', ' ').title())
            kc.fill = _fill(bg); kc.font = _font(color="222222", size=10, bold=True)
            kc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc = ws.cell(row=row, column=2, value=_fmt_value(v))
            vc.fill = _fill(bg); vc.font = _font(color="374151", size=10, bold=False)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            dc = ws.cell(row=row, column=3, value=param_descs.get(k, ''))
            dc.fill = _fill(bg); dc.font = _font(color="6b7280", size=9, bold=False, italic=True)
            dc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row].height = 16
            row += 1
        row += 1

    # ── Annotations ──
    annotations = config.get('annotations', '')
    if annotations:
        ws.merge_cells(f"A{row}:H{row}")
        lbl4 = ws.cell(row=row, column=1, value="NOTAS Y ANOTACIONES")
        lbl4.fill = _fill(secondary)
        lbl4.font = _font(color=fh, size=9)
        lbl4.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 16
        row += 1
        ws.merge_cells(f"A{row}:H{row+2}")
        ac = ws.cell(row=row, column=1, value=annotations)
        ac.fill = _fill("F9FAFB")
        ac.font = _font(color="374151", size=10, bold=False)
        ac.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
        ws.row_dimensions[row].height = 60
        row += 3

    # Column widths
    ws.column_dimensions['A'].width = 36
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20


# ── Tab 2: Escenarios Bear / Base / Bull ───────────────────────────────────────

def _build_scenarios_tab(wb, result: dict, model_name: str, palette: dict):
    ws = wb.create_sheet("Escenarios")
    ws.sheet_view.showGridLines = False

    primary   = palette['primary']
    secondary = palette['secondary']
    light     = palette['light']
    fh        = palette['font_header']

    row = 1
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value=f"ANÁLISIS DE ESCENARIOS — {model_name.upper()}")
    c.fill = _fill(primary)
    c.font = _font(color=fh, size=12, bold=True)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 2

    ws.merge_cells(f"A{row}:G{row}")
    note = ws.cell(row=row, column=1, value=(
        "Bear = ×0.6 (escenario pesimista)  |  Base = ×1.0 (proyección central)  |  "
        "Bull = ×1.4 (escenario optimista)"
    ))
    note.font = _font(color="6b7280", size=9, bold=False, italic=True)
    note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 16
    row += 2

    summary = result.get('summary', {})
    numeric_kpis = [(k, v) for k, v in summary.items()
                    if isinstance(v, (int, float)) and not isinstance(v, bool)]

    if numeric_kpis:
        _write_header_row(ws, row, [
            "Métrica", "Escenario Bear\n(×0.6)", "Escenario Base\n(×1.0)",
            "Escenario Bull\n(×1.4)", "Δ Bear vs Base", "Δ Bull vs Base", "Unidad"
        ], palette, height=32)
        row += 1

        for i, (k, v) in enumerate(numeric_kpis):
            bg = light if i % 2 == 0 else "FFFFFF"
            bear = v * 0.6
            bull = v * 1.4
            d_bear = (bear - v) / abs(v) * 100 if v != 0 else 0
            d_bull = (bull - v) / abs(v) * 100 if v != 0 else 0

            row_vals = [_kpi_label(k), bear, v, bull, f"{d_bear:+.1f}%", f"{d_bull:+.1f}%", _kpi_unit(k)]
            for col, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.fill = _fill(bg)
                if col == 1:
                    cell.font = _font(color="222222", size=10, bold=True)
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif col == 2:   # bear
                    cell.font = _font(color="991b1b", size=10, bold=False)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00' if abs(val) < 10 else '#,##0'
                elif col == 3:   # base
                    cell.font = _font(color="1e40af", size=10, bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00' if abs(val) < 10 else '#,##0'
                elif col == 4:   # bull
                    cell.font = _font(color="065f46", size=10, bold=False)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00' if abs(val) < 10 else '#,##0'
                else:
                    cell.font = _font(color="374151", size=10, bold=False, italic=(col in (5, 6)))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 18
            row += 1

    ws.column_dimensions['A'].width = 34
    for col_letter in ['B', 'C', 'D']:
        ws.column_dimensions[col_letter].width = 20
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 10


# ── Tab 3: Datos Detallados ────────────────────────────────────────────────────

def _build_data_tab(wb, result: dict, model_name: str, palette: dict):
    ws = wb.create_sheet("Datos Detallados")
    ws.sheet_view.showGridLines = False

    primary = palette['primary']
    light   = palette['light']
    accent  = palette['accent']
    fh      = palette['font_header']

    row2 = 1
    ws.merge_cells(f"A{row2}:M{row2}")
    title2 = ws.cell(row=row2, column=1, value=f"DATOS DETALLADOS — {model_name.upper()}")
    title2.fill = _fill(primary)
    title2.font = _font(color=fh, size=12, bold=True)
    title2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row2].height = 26
    row2 += 2

    # Find main data list
    data_list = None
    for key in ('weekly', 'weeks', 'items', 'cohorts', 'segments', 'tiers', 'zones',
                'channels', 'restaurants', 'couriers', 'campaigns', 'periods', 'by_campaign',
                'by_restaurant', 'by_zone'):
        val = result.get(key)
        if isinstance(val, list) and len(val) > 0:
            data_list = val
            break

    if data_list is None:
        for k, v in result.items():
            if isinstance(v, list) and len(v) > 0 and k not in ('summary',):
                data_list = v
                break

    if data_list and isinstance(data_list[0], dict):
        # Flatten nested dicts into separate columns (e.g. by_channel: {A: 1, B: 2} → by_channel_A, by_channel_B)
        flat_headers = []
        for hk, hv in data_list[0].items():
            if isinstance(hv, dict):
                for sub_k in hv.keys():
                    flat_headers.append((hk, sub_k, f"{hk}_{sub_k}"))
            else:
                flat_headers.append((hk, None, hk))

        display_names = [
            (f"{col_k.replace('_', ' ').title()} — {sub_k.replace('_', ' ').title()}" if sub_k else col_k.replace('_', ' ').title())
            for col_k, sub_k, _ in flat_headers
        ]
        _write_header_row(ws, row2, display_names, palette, height=24)
        ws.freeze_panes = f"A{row2 + 1}"
        row2 += 1
        for i, item in enumerate(data_list):
            bg = light if i % 2 == 0 else "FFFFFF"
            is_accent = (i == 0 or i == len(data_list) - 1)
            for col, (hk, sub_k, _) in enumerate(flat_headers, start=1):
                raw = item.get(hk)
                val = raw.get(sub_k) if isinstance(raw, dict) and sub_k else raw
                # Final safety: convert any remaining non-scalar to string
                if isinstance(val, (dict, list)):
                    val = str(val)
                cell = ws.cell(row=row2, column=col, value=val)
                cell.fill = _fill(accent + "22") if is_accent else _fill(bg)
                cell.font = _font(color="222222", size=10, bold=is_accent)
                cell.alignment = Alignment(horizontal="center" if isinstance(val, (int, float)) else "left",
                                           vertical="center", indent=0 if isinstance(val, (int, float)) else 1)
                if isinstance(val, float):
                    if abs(val) >= 1000:
                        cell.number_format = '#,##0'
                    elif 'pct' in (hk if sub_k is None else sub_k) or 'rate' in (hk if sub_k is None else sub_k):
                        cell.number_format = '0.00%' if abs(val) <= 1 else '#,##0.0'
                    else:
                        cell.number_format = '#,##0.00'
            ws.row_dimensions[row2].height = 16
            row2 += 1
    else:
        no_data = ws.cell(row=row2, column=1, value="No hay datos detallados disponibles para este modelo.")
        no_data.font = _font(color="6b7280", size=10, bold=False)

    _auto_width(ws)


# ── Public entry point ─────────────────────────────────────────────────────────

def generate_generic_excel(
    result: dict,
    model_name: str,
    model_id: str,
    config: dict,
    palette_key: str = 'navy',
) -> bytes:
    """
    Generate a 3-tab MBB-quality Excel workbook for any wizard model result.
    Returns bytes suitable for HTTP response.
    """
    palette = get_palette(palette_key)

    wb = Workbook()
    wb.remove(wb.active)

    _build_executive_tab(wb, result, model_name, model_id, config, palette)
    _build_scenarios_tab(wb, result, model_name, palette)
    _build_data_tab(wb, result, model_name, palette)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
