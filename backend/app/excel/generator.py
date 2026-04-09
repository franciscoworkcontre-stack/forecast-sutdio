"""
McKinsey-style Excel report generator using openpyxl.

Tabs generated:
  1. Cover
  2. Executive Summary
  3. Weekly Detail
  4. Monthly Rollup
  5. Cohort Waterfall (if A2)
  6. Promo Detail (if A1)
  7. Scenarios (if B2)
  8. Unit Economics (if B1)
  9. Assumptions Log
"""
import io
from datetime import datetime
from typing import Optional, List

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from ..models.schemas import ForecastResult, ForecastRequest
from .styles import (
    apply_header_style, apply_data_cell, apply_title_style, set_column_widths,
    navy_fill, light_gray_fill, white_fill, thin_border,
    header_font, body_font, title_font, center_align, left_align, right_align,
    FMT_ORDERS, FMT_PCT, FMT_MONEY, FMT_FACTOR, FMT_ORDERS_DEC, FMT_PCT2,
    NAVY, GREEN, RED, AMBER, LIGHT_BLUE, DARK_GRAY
)


def generate_excel(result: ForecastResult, request: ForecastRequest) -> bytes:
    """Generate a McKinsey-style Excel workbook and return as bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _create_cover(wb, result, request)
    _create_executive_summary(wb, result)
    _create_weekly_detail(wb, result, request)
    _create_monthly_rollup(wb, result)

    active = set(request.models_active) if request.models_active else set()

    if "A2" in active and result.cohort_matrix:
        _create_cohort_waterfall(wb, result, request)

    if "A1" in active and result.promo_detail:
        _create_promo_detail(wb, result)

    if "B2" in active and result.scenarios:
        _create_scenarios(wb, result)

    if "B1" in active and result.unit_economics_by_week:
        _create_unit_economics(wb, result)

    _create_assumptions_log(wb, result, request)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def _write_title_bar(ws, row, title, col_span=10):
    ws.row_dimensions[row].height = 28
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


def _write_section_header(ws, row, text, col_span=10):
    ws.row_dimensions[row].height = 18
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


def _create_cover(wb: Workbook, result: ForecastResult, request: ForecastRequest):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 10
    ws.row_dimensions[2].height = 60

    ws.merge_cells("B2:J4")
    cell = ws["B2"]
    cell.value = "Food Delivery Forecast Studio"
    cell.font = Font(name="Calibri", size=28, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B5:J5")
    cell = ws["B5"]
    cell.value = result.name
    cell.font = Font(name="Calibri", size=18, bold=False, color="FFFFFF")
    cell.fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Metadata
    meta_rows = [
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
        ("Country", request.country),
        ("Horizon", f"{request.horizon_weeks} weeks"),
        ("Models Active", ", ".join(request.models_active)),
        ("Total Orders (Forecast)", f"{result.summary.total_orders:,.0f}"),
        ("Peak Week", f"Week {result.summary.peak_week + 1} ({result.summary.peak_orders:,.0f} orders)"),
    ]

    for i, (label, value) in enumerate(meta_rows, start=7):
        ws.cell(row=i, column=2, value=label).font = Font(bold=True, size=11, color="1F4E79")
        ws.cell(row=i, column=4, value=value).font = Font(size=11)

    set_column_widths(ws, {1: 4, 2: 22, 3: 2, 4: 40})


def _create_executive_summary(wb: Workbook, result: ForecastResult):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    _write_title_bar(ws, 1, "Executive Summary", 8)

    # KPI cards row
    kpis = [
        ("Total Orders", f"{result.summary.total_orders:,.0f}", "orders"),
        ("Avg Weekly", f"{result.summary.avg_weekly_orders:,.0f}", "orders/wk"),
        ("Peak Week", f"W{result.summary.peak_week + 1}: {result.summary.peak_orders:,.0f}", "orders"),
        ("Promo Volume", f"{result.summary.promo_total_orders:,.0f}", "incremental orders"),
        ("Promo Cost", f"${result.summary.promo_total_cost:,.0f}", "platform spend"),
        ("ROI", f"{result.summary.promo_roi * 100:.1f}%" if result.summary.promo_roi else "N/A", "promo ROI"),
        ("Weekly CAGR", f"{result.summary.cagr_weekly * 100:.2f}%", "growth/week"),
        ("Payback", f"{result.summary.payback_weeks}w" if result.summary.payback_weeks else "N/A", "weeks to break-even"),
    ]

    # Header
    _write_section_header(ws, 3, "Key Performance Indicators", 8)
    headers = ["Metric", "Value", "Unit"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        apply_header_style(c)

    for row_i, (metric, value, unit) in enumerate(kpis, start=5):
        ws.cell(row=row_i, column=1, value=metric).font = Font(bold=True, size=10)
        ws.cell(row=row_i, column=2, value=value).font = Font(size=10, color="1F4E79")
        ws.cell(row=row_i, column=3, value=unit).font = Font(size=10, color=DARK_GRAY)
        if row_i % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=row_i, column=c).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Weekly summary table
    _write_section_header(ws, 14, "Weekly Order Summary", 8)
    week_headers = ["Week", "Base", "Promo", "Acquisition", "Cannibalization", "Seasonal Factor", "Total Orders", "WoW %"]
    for i, h in enumerate(week_headers, start=1):
        c = ws.cell(row=15, column=i, value=h)
        apply_header_style(c)

    for row_i, w in enumerate(result.weeks[:20], start=16):
        ws.cell(row=row_i, column=1, value=f"W{w.week + 1}").number_format = "@"
        ws.cell(row=row_i, column=2, value=round(w.base_orders)).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=3, value=round(w.promo_orders)).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=4, value=round(w.acquisition_orders)).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=5, value=round(w.cannibalization)).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=6, value=round(w.seasonal_factor, 3)).number_format = "0.000"
        ws.cell(row=row_i, column=7, value=round(w.total_orders)).number_format = FMT_ORDERS
        wow = w.wow_pct / 100 if w.wow_pct is not None else ""
        ws.cell(row=row_i, column=8, value=wow).number_format = FMT_PCT if wow != "" else "@"
        if row_i % 2 == 0:
            for c in range(1, 9):
                ws.cell(row=row_i, column=c).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    set_column_widths(ws, {1: 8, 2: 14, 3: 12, 4: 16, 5: 18, 6: 16, 7: 14, 8: 10})
    ws.freeze_panes = "A16"


def _create_weekly_detail(wb: Workbook, result: ForecastResult, request: ForecastRequest):
    ws = wb.create_sheet("Weekly Detail")
    _write_title_bar(ws, 1, f"Weekly Detail — {result.name}", 10)

    headers = ["Week", "Base Orders", "Promo Orders", "Acq Orders", "Cannibalization",
               "Expansion", "Seasonal Factor", "Mkt Share Factor", "Total Orders", "WoW %"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        apply_header_style(c)

    for row_i, w in enumerate(result.weeks, start=3):
        vals = [
            f"W{w.week + 1}",
            round(w.base_orders),
            round(w.promo_orders),
            round(w.acquisition_orders),
            round(w.cannibalization),
            0,  # expansion — would come from expanded result
            round(w.seasonal_factor, 4),
            round(w.market_share_factor, 4),
            round(w.total_orders),
            (w.wow_pct / 100) if w.wow_pct is not None else "",
        ]
        fmts = ["@", FMT_ORDERS, FMT_ORDERS, FMT_ORDERS, FMT_ORDERS, FMT_ORDERS,
                "0.0000", "0.0000", FMT_ORDERS, FMT_PCT]

        for col_i, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            if fmt and v != "":
                cell.number_format = fmt
            apply_data_cell(cell, row_i)
            cell.alignment = right_align() if col_i > 1 else left_align()

    # Conditional formatting on total orders column (col 9)
    max_row = 2 + len(result.weeks)
    ws.conditional_formatting.add(
        f"I3:I{max_row}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       mid_type="percentile", mid_value=50, mid_color=LIGHT_BLUE,
                       end_type="max", end_color="1F4E79")
    )

    set_column_widths(ws, {1: 8, 2: 14, 3: 14, 4: 14, 5: 16, 6: 12, 7: 16, 8: 18, 9: 14, 10: 10})
    ws.freeze_panes = "B3"


def _create_monthly_rollup(wb: Workbook, result: ForecastResult):
    ws = wb.create_sheet("Monthly Rollup")
    _write_title_bar(ws, 1, "Monthly Rollup", 7)

    headers = ["Month", "Weeks", "Total Orders", "Avg Weekly", "Promo Orders", "Acq Orders", "MoM %"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        apply_header_style(c)

    # Group weeks into months (approx 4 weeks per month)
    months = []
    for month_idx in range(0, len(result.weeks), 4):
        month_weeks = result.weeks[month_idx:month_idx + 4]
        if not month_weeks:
            break
        months.append({
            "month": month_idx // 4 + 1,
            "count": len(month_weeks),
            "total": sum(w.total_orders for w in month_weeks),
            "avg": sum(w.total_orders for w in month_weeks) / len(month_weeks),
            "promo": sum(w.promo_orders for w in month_weeks),
            "acq": sum(w.acquisition_orders for w in month_weeks),
        })

    prev_total = None
    for row_i, m in enumerate(months, start=3):
        mom = (m["total"] - prev_total) / prev_total if prev_total and prev_total > 0 else ""
        ws.cell(row=row_i, column=1, value=f"Month {m['month']}")
        ws.cell(row=row_i, column=2, value=m["count"]).number_format = "0"
        ws.cell(row=row_i, column=3, value=round(m["total"])).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=4, value=round(m["avg"])).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=5, value=round(m["promo"])).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=6, value=round(m["acq"])).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=7, value=mom if mom != "" else "").number_format = FMT_PCT if mom != "" else "@"
        for col in range(1, 8):
            cell = ws.cell(row=row_i, column=col)
            apply_data_cell(cell, row_i)
        prev_total = m["total"]

    set_column_widths(ws, {1: 12, 2: 8, 3: 14, 4: 14, 5: 14, 6: 14, 7: 10})


def _create_cohort_waterfall(wb: Workbook, result: ForecastResult, request: ForecastRequest):
    ws = wb.create_sheet("Cohort Waterfall")
    _write_title_bar(ws, 1, "Restaurant Cohort Waterfall", 14)

    matrix = result.cohort_matrix
    if not matrix:
        return

    horizon = len(result.weeks)
    num_cohorts = len(matrix)

    # Header row: Week labels
    ws.cell(row=2, column=1, value="Cohort \\ Week").font = Font(bold=True)
    for w in range(horizon):
        cell = ws.cell(row=2, column=w + 2, value=f"W{w + 1}")
        apply_header_style(cell, "1F4E79")

    # Cohort rows
    for c_idx, cohort_row in enumerate(matrix):
        row_i = c_idx + 3
        ws.cell(row=row_i, column=1, value=f"Cohort W{c_idx + 1}")
        ws.cell(row=row_i, column=1).font = Font(bold=True, size=9)

        for w_idx, val in enumerate(cohort_row):
            cell = ws.cell(row=row_i, column=w_idx + 2, value=round(val) if val > 0 else "")
            if val > 0:
                cell.number_format = FMT_ORDERS

    # Total row
    total_row = num_cohorts + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, size=10)
    for w_idx in range(horizon):
        total = sum(matrix[c][w_idx] for c in range(num_cohorts) if w_idx < len(matrix[c]))
        cell = ws.cell(row=total_row, column=w_idx + 2, value=round(total))
        cell.number_format = FMT_ORDERS
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    # Conditional color scale on the matrix
    ws.conditional_formatting.add(
        f"B3:${get_column_letter(horizon + 1)}{num_cohorts + 2}",
        ColorScaleRule(start_type="min", start_color="FFFFFF",
                       end_type="max", end_color="1F4E79")
    )

    # Column widths
    ws.column_dimensions["A"].width = 14
    for w in range(horizon):
        ws.column_dimensions[get_column_letter(w + 2)].width = 9

    ws.freeze_panes = "B3"


def _create_promo_detail(wb: Workbook, result: ForecastResult):
    ws = wb.create_sheet("Promo Detail")
    _write_title_bar(ws, 1, "Promotion Detail", 8)

    row = 3
    for promo_data in result.promo_detail:
        ws.cell(row=row, column=1, value=f"Promo #{promo_data['promo_index'] + 1}: {promo_data['promo_type']}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="1F4E79")
        row += 1

        headers = ["Week", "Base Orders", "Incremental", "Residual", "Total Uplift", "Platform Cost"]
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=i, value=h)
            apply_header_style(c)
        row += 1

        for w_data in promo_data["weekly"]:
            ws.cell(row=row, column=1, value=f"W{w_data['week'] + 1}")
            ws.cell(row=row, column=2, value=round(w_data["base_orders"])).number_format = FMT_ORDERS
            ws.cell(row=row, column=3, value=round(w_data["incremental"])).number_format = FMT_ORDERS
            ws.cell(row=row, column=4, value=round(w_data["residual"])).number_format = FMT_ORDERS
            ws.cell(row=row, column=5, value=round(w_data["total"])).number_format = FMT_ORDERS
            ws.cell(row=row, column=6, value=round(w_data["platform_cost"], 2)).number_format = FMT_MONEY
            for col in range(1, 7):
                apply_data_cell(ws.cell(row=row, column=col), row)
            row += 1

        row += 2  # spacing

    set_column_widths(ws, {1: 8, 2: 14, 3: 14, 4: 12, 5: 14, 6: 14})


def _create_scenarios(wb: Workbook, result: ForecastResult):
    ws = wb.create_sheet("Scenarios")
    _write_title_bar(ws, 1, "Scenario Analysis", 10)

    if not result.scenarios:
        return

    # Weekly comparison
    headers = ["Week", "Downside", "Base", "Upside", "Downside vs Base", "Upside vs Base"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        apply_header_style(c)

    base_weeks = result.scenarios.base
    up_weeks = result.scenarios.upside
    down_weeks = result.scenarios.downside

    for row_i, (b, u, d) in enumerate(zip(base_weeks, up_weeks, down_weeks), start=3):
        ws.cell(row=row_i, column=1, value=f"W{b.week + 1}")
        ws.cell(row=row_i, column=2, value=round(d.total_orders)).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=3, value=round(b.total_orders)).number_format = FMT_ORDERS
        ws.cell(row=row_i, column=4, value=round(u.total_orders)).number_format = FMT_ORDERS

        if b.total_orders > 0:
            down_vs = (d.total_orders - b.total_orders) / b.total_orders
            up_vs = (u.total_orders - b.total_orders) / b.total_orders
        else:
            down_vs = 0.0
            up_vs = 0.0

        cell_down = ws.cell(row=row_i, column=5, value=down_vs)
        cell_down.number_format = FMT_PCT
        cell_down.font = Font(color=RED)

        cell_up = ws.cell(row=row_i, column=6, value=up_vs)
        cell_up.number_format = FMT_PCT
        cell_up.font = Font(color="00B050")

        for col in range(1, 7):
            apply_data_cell(ws.cell(row=row_i, column=col), row_i)

    # Tornado chart table
    if result.scenarios.tornado:
        tornado_row = len(base_weeks) + 5
        ws.cell(row=tornado_row, column=1, value="Sensitivity Tornado").font = Font(bold=True, size=12, color="1F4E79")
        tornado_row += 1

        t_headers = ["Variable", "Downside Delta", "Base Total", "Upside Delta"]
        for i, h in enumerate(t_headers, start=1):
            c = ws.cell(row=tornado_row, column=i, value=h)
            apply_header_style(c)
        tornado_row += 1

        for entry in result.scenarios.tornado:
            ws.cell(row=tornado_row, column=1, value=entry.variable)
            ws.cell(row=tornado_row, column=2, value=round(entry.downside_delta)).number_format = FMT_ORDERS
            ws.cell(row=tornado_row, column=3, value=round(entry.base_value)).number_format = FMT_ORDERS
            ws.cell(row=tornado_row, column=4, value=round(entry.upside_delta)).number_format = FMT_ORDERS
            tornado_row += 1

    set_column_widths(ws, {1: 14, 2: 14, 3: 14, 4: 14, 5: 18, 6: 16})
    ws.freeze_panes = "A3"


def _create_unit_economics(wb: Workbook, result: ForecastResult):
    ws = wb.create_sheet("Unit Economics")
    _write_title_bar(ws, 1, "Unit Economics by Week", 12)

    ue_weeks = result.unit_economics_by_week
    if not ue_weeks:
        return

    headers = ["Week", "Orders", "Gross Rev", "Commission", "Delivery Fee", "Ad Rev",
               "Courier Cost", "Support Cost", "Tech Cost", "Subsidy", "Marketing", "CM", "CM %", "EBITDA"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        apply_header_style(c)

    for row_i, ue in enumerate(ue_weeks, start=3):
        vals = [
            f"W{ue.week + 1}",
            round(ue.total_orders),
            round(ue.gross_revenue, 2),
            round(ue.commission_revenue, 2),
            round(ue.delivery_fee_revenue, 2),
            round(ue.ad_revenue, 2),
            round(ue.courier_cost, 2),
            round(ue.support_cost, 2),
            round(ue.tech_cost, 2),
            round(ue.subsidy_cost, 2),
            round(ue.marketing_cost, 2),
            round(ue.contribution_margin, 2),
            ue.contribution_margin_pct,
            round(ue.ebitda, 2),
        ]
        fmts = ["@", FMT_ORDERS, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY,
                FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_MONEY, FMT_PCT, FMT_MONEY]

        for col_i, (v, fmt) in enumerate(zip(vals, fmts), start=1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            if fmt:
                cell.number_format = fmt
            apply_data_cell(cell, row_i)
            if col_i == 12 or col_i == 14:  # CM and EBITDA
                if isinstance(v, (int, float)) and v < 0:
                    cell.font = Font(color=RED)
                elif isinstance(v, (int, float)) and v > 0:
                    cell.font = Font(color="00B050")

    # Totals row
    total_row = len(ue_weeks) + 3
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=2, value=round(sum(u.total_orders for u in ue_weeks))).number_format = FMT_ORDERS
    ws.cell(row=total_row, column=3, value=round(sum(u.gross_revenue for u in ue_weeks), 2)).number_format = FMT_MONEY
    ws.cell(row=total_row, column=12, value=round(sum(u.contribution_margin for u in ue_weeks), 2)).number_format = FMT_MONEY

    for col in range(1, 15):
        ws.cell(row=total_row, column=col).font = Font(bold=True)
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    set_column_widths(ws, {i: 13 for i in range(1, 15)})
    ws.column_dimensions["A"].width = 8
    ws.freeze_panes = "B3"


def _create_assumptions_log(wb: Workbook, result: ForecastResult, request: ForecastRequest):
    ws = wb.create_sheet("Assumptions Log")
    _write_title_bar(ws, 1, "Assumptions & Configuration Log", 4)

    ws.cell(row=3, column=1, value="Parameter").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Value").font = Font(bold=True)
    ws.cell(row=3, column=3, value="Section").font = Font(bold=True)
    apply_header_style(ws.cell(row=3, column=1))
    apply_header_style(ws.cell(row=3, column=2))
    apply_header_style(ws.cell(row=3, column=3))

    assumptions = [
        ("Forecast Name", result.name, "General"),
        ("Horizon (weeks)", request.horizon_weeks, "General"),
        ("Country", request.country, "General"),
        ("Base Weekly Orders", request.base_orders_weekly, "General"),
        ("Models Active", ", ".join(request.models_active), "General"),
    ]

    if request.promo_inputs:
        for i, p in enumerate(request.promo_inputs):
            assumptions += [
                (f"Promo {i+1} Type", p.promo_type, "Promo (A1)"),
                (f"Promo {i+1} Duration (wks)", p.duration_weeks, "Promo (A1)"),
                (f"Promo {i+1} Uplift %", f"{p.uplift_pct}%", "Promo (A1)"),
                (f"Promo {i+1} Segment", p.target_segment, "Promo (A1)"),
                (f"Promo {i+1} Who Funds", p.who_funds, "Promo (A1)"),
            ]

    if request.acquisition_input:
        a = request.acquisition_input
        assumptions += [
            ("Restaurant Type", a.restaurant_type, "Acquisition (A2)"),
            ("Steady State Orders/wk", a.steady_state_orders, "Acquisition (A2)"),
            ("Churn Rate", f"{a.churn_rate * 100:.1f}%", "Acquisition (A2)"),
            ("Churn Week", a.churn_week, "Acquisition (A2)"),
            ("Onboarding Promo Uplift", f"{(a.onboarding_promo_uplift or 0) * 100:.1f}%", "Acquisition (A2)"),
            ("Onboarding Promo Weeks", a.onboarding_promo_weeks, "Acquisition (A2)"),
        ]

    if request.seasonality_input:
        s = request.seasonality_input
        assumptions += [
            ("Seasonality Country", s.country, "Seasonality (A4)"),
            ("Trend Weekly Growth", f"{(s.trend_weekly_growth or 0) * 100:.3f}%", "Seasonality (A4)"),
            ("Use Holidays", str(s.use_holidays), "Seasonality (A4)"),
            ("Use Rain Season", str(s.use_rain_season), "Seasonality (A4)"),
            ("Use Pay Cycles", str(s.use_pay_cycles), "Seasonality (A4)"),
        ]

    if request.unit_economics:
        ue = request.unit_economics
        assumptions += [
            ("Commission %", f"{ue.commission_pct * 100:.1f}%", "Unit Econ (B1)"),
            ("Avg Order Value", f"${ue.avg_order_value:.2f}", "Unit Econ (B1)"),
            ("Delivery Fee", f"${ue.delivery_fee:.2f}", "Unit Econ (B1)"),
            ("Courier Cost/Order", f"${ue.courier_cost_per_order:.2f}", "Unit Econ (B1)"),
            ("Ad Revenue/Order", f"${ue.ad_revenue_per_order:.2f}", "Unit Econ (B1)"),
        ]

    for row_i, (param, value, section) in enumerate(assumptions, start=4):
        ws.cell(row=row_i, column=1, value=param).font = Font(bold=True, size=10)
        ws.cell(row=row_i, column=2, value=str(value)).font = Font(size=10)
        ws.cell(row=row_i, column=3, value=section).font = Font(size=10, color=DARK_GRAY)
        if row_i % 2 == 0:
            for c in range(1, 4):
                ws.cell(row=row_i, column=c).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    set_column_widths(ws, {1: 30, 2: 30, 3: 18})
