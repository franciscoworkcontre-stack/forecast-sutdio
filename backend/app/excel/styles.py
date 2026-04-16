"""Excel styling constants for MBB-grade reports."""
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter

# Colors
NAVY = "002060"
NAVY_LIGHT = "1F3864"
TEAL = "17375E"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9D9D9"
DARK_GRAY = "7F7F7F"
GREEN = "00B050"
RED = "FF0000"
AMBER = "FFC000"
LIGHT_BLUE = "BDD7EE"
DARK_BLUE = "1F4E79"

# Fonts
def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def mono_font(size=10, bold=False, color="000000"):
    return Font(name="Courier New", size=size, bold=bold, color=color)

def title_font(size=14, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

# Fills
def navy_fill():
    return PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

def teal_fill():
    return PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")

def light_gray_fill():
    return PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

def alternate_row_fill():
    return PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")

def white_fill():
    return PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

def light_blue_fill():
    return PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

# Borders
def thin_border():
    thin = Side(style="thin", color=MID_GRAY)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def bottom_border(color=NAVY):
    thick = Side(style="medium", color=color)
    return Border(bottom=thick)

def top_bottom_border():
    thin = Side(style="thin", color=MID_GRAY)
    return Border(top=thin, bottom=thin)

# Alignments
def center_align():
    return Alignment(horizontal="center", vertical="center", wrap_text=False)

def left_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def wrap_align():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# Number formats
FMT_ORDERS = "#,##0"
FMT_ORDERS_DEC = "#,##0.0"
FMT_PCT = "0.0%"
FMT_PCT2 = "0.00%"
FMT_MONEY = "$#,##0.00"
FMT_MONEY_K = "$#,##0"
FMT_FACTOR = "0.000"
FMT_DATE = "YYYY-MM-DD"

def apply_header_style(cell, color=NAVY):
    cell.font = header_font(color=WHITE)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = center_align()
    cell.border = thin_border()

def apply_title_style(cell):
    cell.font = title_font()
    cell.fill = navy_fill()
    cell.alignment = left_align()

def apply_data_cell(cell, row_idx, number_format=None, bold=False):
    cell.font = body_font(bold=bold)
    if row_idx % 2 == 0:
        cell.fill = light_gray_fill()
    else:
        cell.fill = white_fill()
    cell.border = thin_border()
    cell.alignment = right_align()
    if number_format:
        cell.number_format = number_format

def set_column_widths(ws, widths: dict):
    """widths: {col_letter_or_idx: width}"""
    for col, width in widths.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = width
