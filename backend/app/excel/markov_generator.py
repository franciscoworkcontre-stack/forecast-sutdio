"""
MBB-grade Excel report generator for Markov v3 Forecast Model.
Generates 8 tabs: Cover, Exec Summary, Detalle Semanal, Perfiles de Usuario,
Órdenes por Perfil, Matriz de Transición, Supuestos — Funnel, Log de Supuestos.
"""
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ..models.markov_schemas import MarkovForecastResult, MarkovForecastRequest
from .styles import (
    apply_header_style, apply_data_cell, set_column_widths,
    navy_fill, light_gray_fill, white_fill, thin_border,
    header_font, body_font, center_align, left_align, right_align,
    FMT_ORDERS, FMT_PCT, FMT_FACTOR,
    NAVY, NAVY_LIGHT, GREEN, RED, AMBER, DARK_BLUE, WHITE,
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _heatmap_fill(value: float, max_val: float = 1.0):
    """Return blue PatternFill with intensity proportional to value/max_val."""
    ratio = min(1.0, abs(value) / max(max_val, 0.001))
    # Interpolate from white (FFFFFF) to dark blue (1F4E79)
    r = int(255 - ratio * (255 - 31))
    g = int(255 - ratio * (255 - 78))
    b = int(255 - ratio * (255 - 121))
    hex_color = f"{r:02X}{g:02X}{b:02X}"
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _write_title_bar(ws, row: int, title: str, col_span: int = 16):
    ws.row_dimensions[row].height = 28
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


def _write_section_header(ws, row: int, text: str, col_span: int = 12):
    ws.row_dimensions[row].height = 20
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


def _apply_row_fill(ws, row_idx: int, num_cols: int, start_col: int = 1):
    fill = light_gray_fill() if row_idx % 2 == 0 else white_fill()
    for c in range(start_col, start_col + num_cols):
        ws.cell(row=row_idx, column=c).fill = fill


# ── Tab 1: Cover ──────────────────────────────────────────────────────────────

def _create_cover(wb: Workbook, result: MarkovForecastResult, request: MarkovForecastRequest):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False

    summary = result.summary

    # Row 1 spacer
    ws.row_dimensions[1].height = 10

    # Rows 2-4 merged: title
    ws.row_dimensions[2].height = 60
    ws.merge_cells("B2:J4")
    cell = ws["B2"]
    cell.value = "Modelo Markov + Funnel · D1"
    cell.font = Font(name="Calibri", size=28, bold=True, color=WHITE)
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Row 5: forecast name
    ws.row_dimensions[5].height = 32
    ws.merge_cells("B5:J5")
    cell = ws["B5"]
    cell.value = request.name
    cell.font = Font(name="Calibri", size=18, bold=False, color=WHITE)
    cell.fill = PatternFill(start_color=NAVY_LIGHT, end_color=NAVY_LIGHT, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Blank row 6
    ws.row_dimensions[6].height = 10

    # Metadata rows starting row 7, col B (label) and col D (value)
    meta_rows = [
        ("Generado", datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
        ("País", request.country),
        ("Horizonte", f"{request.horizon_weeks} semanas"),
        ("AOV", f"{request.currency} {request.aov:,.0f}"),
        ("Take Rate", f"{request.take_rate * 100:.1f}%"),
        ("Total Órdenes", f"{summary.get('total_orders', 0):,.0f}"),
        ("GMV Total", f"{request.currency} {summary.get('total_gmv', 0):,.0f}"),
        ("CM Total", f"{request.currency} {summary.get('total_contribution_dollar', 0):,.0f}"),
        ("CM Promedio", f"{summary.get('avg_contribution_pct', 0) * 100:.1f}%"),
    ]

    for i, (label, value) in enumerate(meta_rows, start=7):
        ws.row_dimensions[i].height = 18
        lbl = ws.cell(row=i, column=2, value=label)
        lbl.font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
        lbl.alignment = left_align()
        val = ws.cell(row=i, column=4, value=value)
        val.font = Font(name="Calibri", size=11)
        val.alignment = left_align()

    set_column_widths(ws, {1: 4, 2: 22, 3: 2, 4: 40})


# ── Tab 2: Exec Summary ───────────────────────────────────────────────────────

def _create_exec_summary(wb: Workbook, result: MarkovForecastResult, request: MarkovForecastRequest):
    ws = wb.create_sheet("Exec Summary")
    ws.sheet_view.showGridLines = False

    summary = result.summary
    cur = request.currency

    # Title bar row 1
    _write_title_bar(ws, 1, "Resumen Ejecutivo — Markov v3", col_span=12)

    # Blank row 2
    ws.row_dimensions[2].height = 8

    # Section header row 3
    _write_section_header(ws, 3, "KPIs Principales", col_span=12)

    # Headers row 4
    ws.row_dimensions[4].height = 18
    for col, hdr in enumerate(["Métrica", "Valor", "Unidad"], start=1):
        c = ws.cell(row=4, column=col, value=hdr)
        apply_header_style(c, NAVY)

    # KPI data rows 5-14
    def fmt_cur(v): return f"{cur} {v:,.0f}"

    kpis = [
        ("Órdenes Totales",             f"{summary.get('total_orders', 0):,.0f}",                                    "órdenes"),
        ("Órdenes/Semana (avg)",         f"{summary.get('avg_orders_per_week', 0):,.0f}",                             "órdenes/sem"),
        ("GMV Total",                    fmt_cur(summary.get('total_gmv', 0)),                                        cur),
        ("Revenue Neto",                 fmt_cur(summary.get('total_net_revenue', 0)),                                cur),
        ("Gasto P2C — Cupones",          fmt_cur(summary.get('total_gasto_cupon', 0)),                               cur),
        ("Gasto P2C — Free Delivery",   fmt_cur(summary.get('total_gasto_ddc', 0)),                                  cur),
        ("Gasto B2C — Promos/Bundle",   fmt_cur(summary.get('total_gasto_bxsy', 0)),                                 cur),
        ("Total Gastos Comerciales",     fmt_cur(summary.get('total_gastos', 0)),                                     cur),
        ("Contribution Margin $",        fmt_cur(summary.get('total_contribution_dollar', 0)),                        cur),
        ("Contribution Margin %",        f"{summary.get('avg_contribution_pct', 0) * 100:.1f}%",                     "% del revenue"),
    ]

    for row_i, (metric, value, unit) in enumerate(kpis, start=5):
        ws.row_dimensions[row_i].height = 17
        lbl = ws.cell(row=row_i, column=1, value=metric)
        lbl.font = Font(name="Calibri", bold=True, size=10)
        lbl.alignment = left_align()
        val = ws.cell(row=row_i, column=2, value=value)
        val.font = Font(name="Calibri", size=10, color=DARK_BLUE)
        val.alignment = right_align()
        unt = ws.cell(row=row_i, column=3, value=unit)
        unt.font = Font(name="Calibri", size=10, color="7F7F7F")
        unt.alignment = left_align()
        if row_i % 2 == 0:
            for col in range(1, 4):
                ws.cell(row=row_i, column=col).fill = light_gray_fill()

    # Blank row 15
    ws.row_dimensions[15].height = 8

    # Section header row 16
    _write_section_header(ws, 16, "Semana Pico y Base", col_span=12)

    # Row 17 headers
    week_headers = ["Semana", "Base", "Incremental", "Total", "GMV", "Revenue", "Gasto", "CM$", "CM%"]
    ws.row_dimensions[17].height = 18
    for col, hdr in enumerate(week_headers, start=1):
        c = ws.cell(row=17, column=col, value=hdr)
        apply_header_style(c, NAVY)

    # Top 3 weeks by orders_total
    sorted_weeks = sorted(result.weeks, key=lambda w: w.orders_total, reverse=True)[:3]
    for row_i, w in enumerate(sorted_weeks, start=18):
        ws.row_dimensions[row_i].height = 17
        vals = [
            f"S{w.week + 1}",
            round(w.orders_base),
            round(w.orders_incremental),
            round(w.orders_total),
            round(w.gmv),
            round(w.net_revenue),
            round(w.total_gastos),
            round(w.contribution_dollar),
            f"{w.contribution_pct * 100:.1f}%",
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col, value=v)
            cell.font = body_font()
            cell.alignment = right_align() if col > 1 else left_align()
            if row_i % 2 == 0:
                cell.fill = light_gray_fill()

    set_column_widths(ws, {1: 28, 2: 18, 3: 18, 4: 15})


# ── Tab 3: Detalle Semanal ────────────────────────────────────────────────────

def _create_weekly_detail(wb: Workbook, result: MarkovForecastResult, request: MarkovForecastRequest):
    ws = wb.create_sheet("Detalle Semanal")
    ws.sheet_view.showGridLines = False

    _write_title_bar(ws, 1, "Detalle Semanal de Órdenes y P&L", col_span=16)

    # Headers row 2
    headers = [
        "Semana", "Órdenes Base", "Órd. Incremental", "Órd. Total", "WoW%",
        "GMV", "Revenue Neto", "Gasto P2C Cupones", "Gasto P2C DDC", "Gasto B2C BxSy",
        "Total Gastos", "CM$", "CM%", "Costo/Orden", "Traffic Mult", "Conv Mult",
    ]
    ws.row_dimensions[2].height = 18
    for col, hdr in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        apply_header_style(c, NAVY)

    # Freeze panes after headers
    ws.freeze_panes = ws.cell(row=3, column=1)

    weeks = result.weeks
    for row_i, w in enumerate(weeks, start=3):
        ws.row_dimensions[row_i].height = 16

        # WoW%
        prev = weeks[row_i - 4] if row_i - 4 >= 0 else None  # row_i starts at 3, weeks[0]=first
        prev_idx = row_i - 3 - 1  # index into weeks
        if prev_idx > 0 and weeks[prev_idx - 1].orders_total > 0:
            wow = (w.orders_total - weeks[prev_idx - 1].orders_total) / weeks[prev_idx - 1].orders_total
        else:
            wow = None

        vals = [
            (f"S{w.week + 1}",           "@",        None),
            (w.orders_base,               FMT_ORDERS, None),
            (w.orders_incremental,        FMT_ORDERS, None),
            (w.orders_total,              FMT_ORDERS, None),
            (wow,                         FMT_PCT,    None),
            (w.gmv,                       "#,##0.00", None),
            (w.net_revenue,               "#,##0.00", None),
            (w.gasto_cupon,               "#,##0.00", None),
            (w.gasto_ddc,                 "#,##0.00", None),
            (w.gasto_bxsy,                "#,##0.00", None),
            (w.total_gastos,              "#,##0.00", None),
            (w.contribution_dollar,       "#,##0.00", None),
            (w.contribution_pct,          FMT_PCT,    None),
            (w.cost_per_order,            "#,##0.00", None),
            (w.traffic_mult,              "0.000",    None),
            (w.conv_mult,                 "0.000",    None),
        ]

        for col, (v, fmt, _) in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col, value=v)
            apply_data_cell(cell, row_i)
            if v is not None and fmt and fmt != "@":
                cell.number_format = fmt
            if col == 1:
                cell.alignment = left_align()
            else:
                cell.alignment = right_align()

            # WoW% coloring (col 5)
            if col == 5 and wow is not None:
                cell.font = Font(name="Calibri", size=10, color=GREEN if wow >= 0 else RED)

            # CM% coloring (col 13)
            if col == 13:
                cm = w.contribution_pct
                if cm >= 0.20:
                    cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, color=WHITE)
                elif cm >= 0.05:
                    cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, color="000000")
                else:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(name="Calibri", size=10, color=WHITE)

    # Totals row
    total_row = 3 + len(weeks)
    ws.row_dimensions[total_row].height = 18
    total_label = ws.cell(row=total_row, column=1, value="TOTAL")
    total_label.font = Font(name="Calibri", bold=True, size=10)
    total_label.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_label.alignment = left_align()

    total_vals = [
        (2,  sum(w.orders_base for w in weeks),          FMT_ORDERS),
        (3,  sum(w.orders_incremental for w in weeks),   FMT_ORDERS),
        (4,  sum(w.orders_total for w in weeks),         FMT_ORDERS),
        (5,  None,                                        "@"),   # WoW% N/A for total
        (6,  sum(w.gmv for w in weeks),                  "#,##0.00"),
        (7,  sum(w.net_revenue for w in weeks),          "#,##0.00"),
        (8,  sum(w.gasto_cupon for w in weeks),          "#,##0.00"),
        (9,  sum(w.gasto_ddc for w in weeks),            "#,##0.00"),
        (10, sum(w.gasto_bxsy for w in weeks),           "#,##0.00"),
        (11, sum(w.total_gastos for w in weeks),         "#,##0.00"),
        (12, sum(w.contribution_dollar for w in weeks),  "#,##0.00"),
        (13, sum(w.contribution_pct for w in weeks) / len(weeks) if weeks else 0, FMT_PCT),
        (14, sum(w.cost_per_order for w in weeks) / len(weeks) if weeks else 0, "#,##0.00"),
        (15, sum(w.traffic_mult for w in weeks) / len(weeks) if weeks else 0, "0.000"),
        (16, sum(w.conv_mult for w in weeks) / len(weeks) if weeks else 0, "0.000"),
    ]
    for col, v, fmt in total_vals:
        cell = ws.cell(row=total_row, column=col, value=v)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        cell.alignment = right_align()
        if v is not None and fmt != "@":
            cell.number_format = fmt

    set_column_widths(ws, {
        1: 8, 2: 16, 3: 18, 4: 14, 5: 10,
        6: 16, 7: 16, 8: 14, 9: 14, 10: 14,
        11: 14, 12: 14, 13: 10, 14: 14, 15: 14, 16: 14,
    })


# ── Tab 4: Perfiles de Usuario ────────────────────────────────────────────────

def _create_user_profiles(wb: Workbook, result: MarkovForecastResult, request: MarkovForecastRequest):
    ws = wb.create_sheet("Perfiles de Usuario")
    ws.sheet_view.showGridLines = False

    profiles = request.profiles
    n = len(profiles)
    col_span = 2 + n

    _write_title_bar(ws, 1, "Evolución de Usuarios por Perfil (Cadena de Markov)", col_span=col_span)

    # Section header row 2
    ws.row_dimensions[2].height = 18
    cell = ws.cell(row=2, column=1, value="Usuarios en cada estado al inicio de cada semana")
    cell.font = Font(name="Calibri", size=10, italic=True, color="7F7F7F")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)

    # Headers row 3
    ws.row_dimensions[3].height = 18
    hdr_cells = ["Semana"] + [p.name for p in profiles] + ["Total Usuarios"]
    for col, hdr in enumerate(hdr_cells, start=1):
        c = ws.cell(row=3, column=col, value=hdr)
        apply_header_style(c, NAVY)

    ws.freeze_panes = ws.cell(row=4, column=1)

    # Data rows
    weeks = result.weeks
    for row_i, w in enumerate(weeks, start=4):
        ws.row_dimensions[row_i].height = 16
        cell = ws.cell(row=row_i, column=1, value=f"S{w.week + 1}")
        apply_data_cell(cell, row_i)
        cell.alignment = left_align()

        total = 0.0
        for col, p in enumerate(profiles, start=2):
            v = w.profile_users.get(p.id, 0.0)
            total += v
            c = ws.cell(row=row_i, column=col, value=round(v))
            apply_data_cell(c, row_i)
            c.number_format = FMT_ORDERS
            c.alignment = right_align()

        c_total = ws.cell(row=row_i, column=2 + n, value=round(total))
        apply_data_cell(c_total, row_i)
        c_total.number_format = FMT_ORDERS
        c_total.alignment = right_align()
        c_total.font = Font(name="Calibri", bold=True, size=10)

    # Totals row
    total_row = 4 + len(weeks)
    ws.row_dimensions[total_row].height = 18
    cell = ws.cell(row=total_row, column=1, value="TOTAL")
    cell.font = Font(name="Calibri", bold=True, size=10)
    cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    cell.alignment = left_align()

    grand_total = 0.0
    for col, p in enumerate(profiles, start=2):
        s = sum(w.profile_users.get(p.id, 0.0) for w in weeks)
        grand_total += s
        c = ws.cell(row=total_row, column=col, value=round(s))
        c.font = Font(name="Calibri", bold=True, size=10)
        c.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        c.number_format = FMT_ORDERS
        c.alignment = right_align()

    c_gt = ws.cell(row=total_row, column=2 + n, value=round(grand_total))
    c_gt.font = Font(name="Calibri", bold=True, size=10)
    c_gt.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    c_gt.number_format = FMT_ORDERS
    c_gt.alignment = right_align()

    widths = {1: 8}
    for i in range(n):
        widths[i + 2] = 18
    widths[2 + n] = 16
    set_column_widths(ws, widths)


# ── Tab 5: Órdenes por Perfil ─────────────────────────────────────────────────

def _create_orders_by_profile(wb: Workbook, result: MarkovForecastResult, request: MarkovForecastRequest):
    ws = wb.create_sheet("Órdenes por Perfil")
    ws.sheet_view.showGridLines = False

    profiles = request.profiles
    n = len(profiles)
    num_extra = 3  # Total, Incremental Total, % Incremental
    col_span = 1 + n + num_extra

    _write_title_bar(ws, 1, "Órdenes Semanales por Perfil de Usuario", col_span=col_span)

    # Blank row 2
    ws.row_dimensions[2].height = 8

    # Headers row 3
    ws.row_dimensions[3].height = 18
    hdr_cells = ["Semana"] + [p.name for p in profiles] + ["Total", "Incremental Total", "% Incremental"]
    for col, hdr in enumerate(hdr_cells, start=1):
        c = ws.cell(row=3, column=col, value=hdr)
        apply_header_style(c, NAVY)

    ws.freeze_panes = ws.cell(row=4, column=1)

    weeks = result.weeks
    for row_i, w in enumerate(weeks, start=4):
        ws.row_dimensions[row_i].height = 16
        cell = ws.cell(row=row_i, column=1, value=f"S{w.week + 1}")
        apply_data_cell(cell, row_i)
        cell.alignment = left_align()

        total_orders = 0.0
        for col, p in enumerate(profiles, start=2):
            v = w.profile_orders.get(p.id, 0.0)
            total_orders += v
            c = ws.cell(row=row_i, column=col, value=round(v))
            apply_data_cell(c, row_i)
            c.number_format = FMT_ORDERS
            c.alignment = right_align()

        # Total
        c_total = ws.cell(row=row_i, column=2 + n, value=round(total_orders))
        apply_data_cell(c_total, row_i)
        c_total.number_format = FMT_ORDERS
        c_total.alignment = right_align()
        c_total.font = Font(name="Calibri", bold=True, size=10)

        # Incremental Total
        incr_total = w.orders_incremental
        c_incr = ws.cell(row=row_i, column=3 + n, value=round(incr_total))
        apply_data_cell(c_incr, row_i)
        c_incr.number_format = FMT_ORDERS
        c_incr.alignment = right_align()

        # % Incremental
        pct_incr = incr_total / total_orders if total_orders > 0 else 0.0
        c_pct = ws.cell(row=row_i, column=4 + n, value=pct_incr)
        apply_data_cell(c_pct, row_i)
        c_pct.number_format = FMT_PCT
        c_pct.alignment = right_align()

    # Totals row
    total_row = 4 + len(weeks)
    ws.row_dimensions[total_row].height = 18

    totals_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    label = ws.cell(row=total_row, column=1, value="TOTAL")
    label.font = Font(name="Calibri", bold=True, size=10)
    label.fill = totals_fill
    label.alignment = left_align()

    grand_orders = 0.0
    for col, p in enumerate(profiles, start=2):
        s = sum(w.profile_orders.get(p.id, 0.0) for w in weeks)
        grand_orders += s
        c = ws.cell(row=total_row, column=col, value=round(s))
        c.font = Font(name="Calibri", bold=True, size=10)
        c.fill = totals_fill
        c.number_format = FMT_ORDERS
        c.alignment = right_align()

    grand_total = sum(w.orders_total for w in weeks)
    grand_incr = sum(w.orders_incremental for w in weeks)

    for col, v, fmt in [
        (2 + n, round(grand_total),   FMT_ORDERS),
        (3 + n, round(grand_incr),    FMT_ORDERS),
        (4 + n, grand_incr / grand_total if grand_total > 0 else 0.0, FMT_PCT),
    ]:
        c = ws.cell(row=total_row, column=col, value=v)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.fill = totals_fill
        c.number_format = fmt
        c.alignment = right_align()

    widths = {1: 8}
    for i in range(n):
        widths[i + 2] = 18
    widths[2 + n] = 16
    widths[3 + n] = 18
    widths[4 + n] = 16
    set_column_widths(ws, widths)


# ── Tab 6: Matriz de Transición ───────────────────────────────────────────────

def _create_transition_matrix(wb: Workbook, request: MarkovForecastRequest):
    ws = wb.create_sheet("Matriz de Transición")
    ws.sheet_view.showGridLines = False

    tm = request.transition_matrix
    profile_ids = tm.profile_ids
    matrix = tm.matrix
    n = len(profile_ids)

    # Map profile_id -> name
    id_to_name = {p.id: p.name for p in request.profiles}
    profile_names = [id_to_name.get(pid, pid) for pid in profile_ids]

    col_span = 1 + n + 1  # label + n states + row sum
    _write_title_bar(ws, 1, "Matriz de Transición de Markov (NxN)", col_span=col_span)

    ws.row_dimensions[2].height = 18
    cell = ws.cell(row=2, column=1, value="Probabilidad de transición entre estados en una semana")
    cell.font = Font(name="Calibri", size=10, italic=True, color="7F7F7F")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)

    # Row 3 spacer
    ws.row_dimensions[3].height = 8

    # Row 4: column headers (blank cell in A, then profile names)
    ws.row_dimensions[4].height = 18
    ws.cell(row=4, column=1, value="")
    for col, name in enumerate(profile_names, start=2):
        c = ws.cell(row=4, column=col, value=name)
        apply_header_style(c, NAVY)
    # Row sum header
    c_sum = ws.cell(row=4, column=2 + n, value="Suma Fila")
    apply_header_style(c_sum, NAVY)

    # Matrix rows
    for row_i, (pid, row_data) in enumerate(zip(profile_ids, matrix), start=5):
        ws.row_dimensions[row_i].height = 17

        # Row label
        lbl = ws.cell(row=row_i, column=1, value=id_to_name.get(pid, pid))
        lbl.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
        lbl.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
        lbl.alignment = left_align()

        row_sum = sum(row_data)
        for col_j, val in enumerate(row_data, start=2):
            cell = ws.cell(row=row_i, column=col_j, value=val)
            cell.number_format = "0.0%"
            cell.alignment = center_align()
            cell.border = thin_border()
            # Heatmap fill
            cell.fill = _heatmap_fill(val, max_val=1.0)
            # Font color: white for darker fills, black for lighter
            ratio = min(1.0, val)
            cell.font = Font(
                name="Calibri", size=10,
                color=WHITE if ratio >= 0.4 else "000000"
            )

        # Row sum
        c_sum = ws.cell(row=row_i, column=2 + n, value=row_sum)
        c_sum.number_format = "0.000"
        c_sum.alignment = center_align()
        c_sum.border = thin_border()
        if abs(row_sum - 1.0) < 0.002:
            c_sum.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
            c_sum.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
        else:
            c_sum.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            c_sum.font = Font(name="Calibri", size=10, bold=True, color=WHITE)

    # Column widths
    widths = {1: 22}
    for i in range(n):
        widths[i + 2] = 14
    widths[2 + n] = 14
    set_column_widths(ws, widths)


# ── Tab 7: Supuestos — Funnel ─────────────────────────────────────────────────

def _create_funnel_assumptions(wb: Workbook, request: MarkovForecastRequest):
    ws = wb.create_sheet("Supuestos — Funnel")
    ws.sheet_view.showGridLines = False

    _write_title_bar(ws, 1, "Parámetros del Funnel por Perfil", col_span=14)

    # Funnel parameters headers
    funnel_headers = [
        "Perfil", "Open App%", "Sesiones/sem", "Ve Vertical%",
        "Entry Topic", "Entry Feed", "Entry Filter",
        "P1 Topic", "P1 Feed", "P1 Filter",
        "P2 Topic", "P2 Feed", "P2 Filter",
    ]
    ws.row_dimensions[2].height = 18
    for col, hdr in enumerate(funnel_headers, start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        apply_header_style(c, NAVY)

    # Map profile_id -> name
    id_to_name = {p.id: p.name for p in request.profiles}

    for row_i, fp in enumerate(request.funnel_params, start=3):
        ws.row_dimensions[row_i].height = 16
        vals = [
            id_to_name.get(fp.profile_id, fp.profile_id),
            fp.open_app_pct,
            fp.avg_weekly_sessions,
            fp.see_vertical_pct,
            fp.entry_topic,
            fp.entry_feed,
            fp.entry_filter,
            fp.p1_topic,
            fp.p1_feed,
            fp.p1_filter,
            fp.p2_topic,
            fp.p2_feed,
            fp.p2_filter,
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=row_i, column=col, value=v)
            apply_data_cell(cell, row_i)
            if col == 1:
                cell.alignment = left_align()
                cell.font = Font(name="Calibri", bold=True, size=10)
            else:
                cell.number_format = "0.0%"
                cell.alignment = right_align()

    # Blank rows
    spacer_row = 3 + len(request.funnel_params) + 1
    ws.row_dimensions[spacer_row].height = 12

    # Active levers section
    lever_header_row = spacer_row + 1
    _write_section_header(ws, lever_header_row, "Levers Activos", col_span=14)

    lever_col_row = lever_header_row + 1
    ws.row_dimensions[lever_col_row].height = 18
    for col, hdr in enumerate(["ID", "Nombre", "Tipo", "Uplift Base", "Estado"], start=1):
        c = ws.cell(row=lever_col_row, column=col, value=hdr)
        apply_header_style(c, NAVY)

    for row_i, lever in enumerate(request.levers, start=lever_col_row + 1):
        ws.row_dimensions[row_i].height = 16
        status = "ACTIVO" if lever.active else "INACTIVO"
        status_color = GREEN if lever.active else "7F7F7F"
        lvls = [lever.id, lever.name, lever.lever_type, lever.base_uplift, status]
        for col, v in enumerate(lvls, start=1):
            cell = ws.cell(row=row_i, column=col, value=v)
            apply_data_cell(cell, row_i)
            if col == 4:
                cell.number_format = "0.0%"
                cell.alignment = right_align()
            elif col == 5:
                cell.font = Font(name="Calibri", size=10, bold=True, color=status_color)
                cell.alignment = center_align()
            elif col == 1:
                cell.alignment = left_align()
            else:
                cell.alignment = left_align()

    widths = {1: 18}
    for i in range(1, 14):
        widths[i + 1] = 14
    set_column_widths(ws, widths)


# ── Tab 8: Log de Supuestos ───────────────────────────────────────────────────

def _create_assumptions_log(wb: Workbook, result: MarkovForecastResult, request: MarkovForecastRequest):
    ws = wb.create_sheet("Log de Supuestos")
    ws.sheet_view.showGridLines = False

    _write_title_bar(ws, 1, "Log de Supuestos — Auditoría Completa", col_span=4)

    # Headers
    ws.row_dimensions[2].height = 18
    for col, hdr in enumerate(["Parámetro", "Valor"], start=1):
        c = ws.cell(row=2, column=col, value=hdr)
        apply_header_style(c, NAVY)

    rows = []

    # General config
    rows += [
        ("Nombre del Forecast", request.name),
        ("País", request.country),
        ("Horizonte (semanas)", request.horizon_weeks),
        ("AOV", request.aov),
        ("Take Rate", request.take_rate),
        ("Moneda", request.currency),
        ("Overlap Factor", request.overlap_factor),
        (None, None),  # blank row
    ]

    # Profiles
    rows.append(("PERFILES DE USUARIO", "__section__"))
    for p in request.profiles:
        rows.append((f"  {p.id} — {p.name}", p.initial_users))
    rows.append((None, None))

    # Ramp config
    rows.append(("RAMP CONFIG", "__section__"))
    rows.append(("  Semanas de ramp-up", request.ramp_config.ramp_weeks))
    rows.append(("  Tipo de curva", request.ramp_config.curve_type))
    rows.append((None, None))

    # Acquisition
    rows.append(("ADQUISICIÓN", "__section__"))
    rows.append(("  Activo", str(request.acquisition.active)))
    rows.append(("  EFO Share", request.acquisition.efo_share))
    rows.append(("  Alpha", request.acquisition.alpha))
    rows.append(("  New User Orders Ratio", request.acquisition.new_user_orders_ratio))
    rows.append(("  WoW Cap", request.acquisition.wow_cap))
    rows.append((None, None))

    # Levers
    rows.append(("LEVERS", "__section__"))
    for lever in request.levers:
        rows.append((
            f"  {lever.id} ({lever.lever_type})",
            f"{lever.base_uplift * 100:.1f}% uplift — {'ACTIVO' if lever.active else 'inactivo'}"
        ))
    rows.append((None, None))

    # Generated
    rows.append(("GENERADO", datetime.now().strftime("%Y-%m-%d %H:%M UTC")))

    for row_i, (param, value) in enumerate(rows, start=3):
        ws.row_dimensions[row_i].height = 16

        if param is None:
            # blank row
            continue

        if value == "__section__":
            # Section header in bold navy
            cell = ws.cell(row=row_i, column=1, value=param)
            cell.font = Font(name="Calibri", bold=True, size=11, color=DARK_BLUE)
            cell.alignment = left_align()
        else:
            p_cell = ws.cell(row=row_i, column=1, value=param)
            p_cell.font = Font(name="Calibri", bold=param and not param.startswith("  "), size=10)
            p_cell.alignment = left_align()
            p_cell.border = thin_border()

            v_cell = ws.cell(row=row_i, column=2, value=str(value) if not isinstance(value, str) else value)
            v_cell.font = Font(name="Calibri", size=10)
            v_cell.alignment = left_align()
            v_cell.border = thin_border()

            if row_i % 2 == 0:
                p_cell.fill = light_gray_fill()
                v_cell.fill = light_gray_fill()

    set_column_widths(ws, {1: 40, 2: 30})


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_markov_excel(result: MarkovForecastResult, request: MarkovForecastRequest, palette_key: str = 'navy') -> bytes:
    """Generate an MBB-grade Excel workbook for the Markov v3 forecast and return as bytes."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _create_cover(wb, result, request)
    _create_exec_summary(wb, result, request)
    _create_weekly_detail(wb, result, request)
    _create_user_profiles(wb, result, request)
    _create_orders_by_profile(wb, result, request)
    _create_transition_matrix(wb, request)
    _create_funnel_assumptions(wb, request)
    _create_assumptions_log(wb, result, request)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
